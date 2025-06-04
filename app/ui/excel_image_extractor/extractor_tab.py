# -*- coding: utf-8 -*-
"""Excel图片提取选项卡模块"""

import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import openpyxl
from openpyxl.drawing.image import Image
import io
from PIL import Image as PILImage
import string
import threading

# 设置PIL图像大小限制，避免解压炸弹攻击的限制
PILImage.MAX_IMAGE_PIXELS = None

class ExtractorTab:
    """Excel图片提取选项卡"""
    
    def __init__(self, parent, notebook, theme, config):
        self.parent = parent
        self.notebook = notebook
        self.theme = theme
        self.config = config
        self.update_status = None  # 状态更新回调函数
        
        # 创建选项卡
        self.frame = ttk.Frame(notebook)
        notebook.add(self.frame, text="📊 图片提取")
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置用户界面"""
        # 创建主框架
        main_frame = ttk.Frame(self.frame, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 左侧控制面板
        control_frame = ttk.LabelFrame(main_frame, text="控制面板", padding="10")
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        
        # Excel文件选择
        file_frame = ttk.Frame(control_frame)
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(file_frame, text="Excel文件：").pack(anchor=tk.W)
        
        file_select_frame = ttk.Frame(file_frame)
        file_select_frame.pack(fill=tk.X, pady=5)
        
        self.file_path_var = tk.StringVar()
        ttk.Entry(file_select_frame, textvariable=self.file_path_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(file_select_frame, text="浏览", command=self.browse_file).pack(side=tk.RIGHT, padx=(5, 0))
        
        # 工作表选择
        worksheet_frame = ttk.Frame(control_frame)
        worksheet_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(worksheet_frame, text="工作表：").pack(side=tk.LEFT, padx=5)
        self.worksheet_var = tk.StringVar()
        self.worksheet_combo = ttk.Combobox(worksheet_frame, textvariable=self.worksheet_var, width=30, state="readonly")
        self.worksheet_combo.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 输出目录选择
        output_frame = ttk.Frame(control_frame)
        output_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(output_frame, text="输出目录：").pack(anchor=tk.W)
        
        output_select_frame = ttk.Frame(output_frame)
        output_select_frame.pack(fill=tk.X, pady=5)
        
        self.output_path_var = tk.StringVar()
        ttk.Entry(output_select_frame, textvariable=self.output_path_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(output_select_frame, text="浏览", command=self.browse_output_folder).pack(side=tk.RIGHT, padx=(5, 0))
        
        # 操作按钮
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="加载工作表", command=self.load_worksheets).pack(fill=tk.X, pady=2)
        ttk.Button(button_frame, text="预览图片", command=self.preview_images).pack(fill=tk.X, pady=2)
        ttk.Button(button_frame, text="提取图片", command=self.extract_images).pack(fill=tk.X, pady=2)
        
        # 右侧预览区域
        preview_frame = ttk.LabelFrame(main_frame, text="预览区域", padding="10")
        preview_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # 创建滚动区域
        canvas = tk.Canvas(preview_frame, bg="white")
        scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.canvas = canvas
    
    def browse_file(self):
        """浏览Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            if self.update_status:
                self.update_status(f"已选择文件: {os.path.basename(file_path)}")
    
    def browse_output_folder(self):
        """浏览输出目录"""
        folder_path = filedialog.askdirectory(title="选择输出目录")
        if folder_path:
            self.output_path_var.set(folder_path)
            if self.update_status:
                self.update_status(f"输出目录: {folder_path}")
    
    def load_worksheets(self):
        """加载工作表列表"""
        file_path = self.file_path_var.get()
        if not file_path:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet_names = workbook.sheetnames
            self.worksheet_combo['values'] = worksheet_names
            if worksheet_names:
                self.worksheet_combo.set(worksheet_names[0])
            workbook.close()
            
            if self.update_status:
                self.update_status(f"已加载 {len(worksheet_names)} 个工作表")
        except Exception as e:
            messagebox.showerror("错误", f"加载工作表失败: {str(e)}")
    
    def preview_images(self):
        """预览图片"""
        file_path = self.file_path_var.get()
        worksheet_name = self.worksheet_var.get()
        
        if not file_path or not worksheet_name:
            messagebox.showwarning("警告", "请先选择Excel文件和工作表")
            return
        
        # 清空预览区域
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook[worksheet_name]
            
            images = []
            for image in worksheet._images:
                images.append(image)
            
            if not images:
                ttk.Label(self.scrollable_frame, text="该工作表中没有找到图片").pack(pady=20)
            else:
                for i, image in enumerate(images):
                    try:
                        # 获取图片数据
                        img_data = image._data()
                        pil_image = PILImage.open(io.BytesIO(img_data))
                        
                        # 调整图片大小用于预览
                        pil_image.thumbnail((200, 200), PILImage.Resampling.LANCZOS)
                        
                        # 转换为Tkinter可用的格式
                        from PIL import ImageTk
                        tk_image = ImageTk.PhotoImage(pil_image)
                        
                        # 创建预览标签
                        frame = ttk.Frame(self.scrollable_frame)
                        frame.pack(pady=5, fill=tk.X)
                        
                        label = tk.Label(frame, image=tk_image)
                        label.image = tk_image  # 保持引用
                        label.pack(side=tk.LEFT, padx=5)
                        
                        info_label = ttk.Label(frame, text=f"图片 {i+1}\n位置: {image.anchor}")
                        info_label.pack(side=tk.LEFT, padx=10)
                        
                    except Exception as e:
                        error_label = ttk.Label(self.scrollable_frame, text=f"图片 {i+1} 预览失败: {str(e)}")
                        error_label.pack(pady=2)
            
            workbook.close()
            
            if self.update_status:
                self.update_status(f"预览完成，找到 {len(images)} 张图片")
                
        except Exception as e:
            messagebox.showerror("错误", f"预览图片失败: {str(e)}")
    
    def extract_images(self):
        """提取图片"""
        file_path = self.file_path_var.get()
        worksheet_name = self.worksheet_var.get()
        output_path = self.output_path_var.get()
        
        if not file_path or not worksheet_name:
            messagebox.showwarning("警告", "请先选择Excel文件和工作表")
            return
        
        if not output_path:
            messagebox.showwarning("警告", "请选择输出目录")
            return
        
        # 在新线程中执行提取操作
        thread = threading.Thread(target=self._extract_images_thread, 
                                 args=(file_path, worksheet_name, output_path))
        thread.daemon = True
        thread.start()
    
    def _extract_images_thread(self, file_path, worksheet_name, output_path):
        """在线程中执行图片提取"""
        try:
            if self.update_status:
                self.update_status("正在提取图片...")
            
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook[worksheet_name]
            
            # 创建输出目录
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_dir = os.path.join(output_path, f"{base_name}_{worksheet_name}_images")
            os.makedirs(output_dir, exist_ok=True)
            
            extracted_count = 0
            for i, image in enumerate(worksheet._images):
                try:
                    # 获取图片数据
                    img_data = image._data()
                    
                    # 确定文件扩展名
                    pil_image = PILImage.open(io.BytesIO(img_data))
                    format_ext = pil_image.format.lower() if pil_image.format else 'png'
                    
                    # 保存图片
                    filename = f"image_{i+1:03d}.{format_ext}"
                    filepath = os.path.join(output_dir, filename)
                    
                    with open(filepath, 'wb') as f:
                        f.write(img_data)
                    
                    extracted_count += 1
                    
                except Exception as e:
                    print(f"提取图片 {i+1} 失败: {str(e)}")
            
            workbook.close()
            
            if self.update_status:
                self.update_status(f"提取完成！成功提取 {extracted_count} 张图片到 {output_dir}")
            
            # 在主线程中显示完成消息
            self.parent.after(0, lambda: messagebox.showinfo(
                "完成", 
                f"图片提取完成！\n成功提取 {extracted_count} 张图片\n保存位置: {output_dir}"
            ))
            
        except Exception as e:
            error_msg = f"提取图片失败: {str(e)}"
            if self.update_status:
                self.update_status(error_msg)
            self.parent.after(0, lambda: messagebox.showerror("错误", error_msg))
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import openpyxl
from openpyxl import Workbook
import os
import tempfile
import json
from ..utils.coordinate_converter import wgs84_to_gcj02, gcj02_to_wgs84, convert_coordinates, calculate_distance
from ..utils.amap_api import amap_api
from ..utils.history_manager import history_manager
from config import config

class AmapTool:
    def __init__(self, parent, theme):
        self.parent = parent
        self.theme = theme
        self.status_text = tk.StringVar(value="就绪")
        
        # 初始化配置和历史记录
        self.config = config
        self.history_manager = history_manager
        
        # 检查API密钥
        self.check_api_key()
        
        self.setup_ui()
    
    def check_api_key(self):
        """检查API密钥配置"""
        api_key = self.config.get_amap_api_key()
        if not api_key:
            self.show_api_key_dialog()
        else:
            amap_api.set_api_key(api_key)
    
    def show_api_key_dialog(self):
        """显示API密钥配置对话框"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("配置高德地图API密钥")
        dialog.geometry("500x300")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # 说明文本
        info_text = """为了使用高德地图的完整功能，请配置您的API密钥：

1. 访问高德开放平台：https://lbs.amap.com/
2. 注册账号并创建应用
3. 获取Web服务API密钥
4. 在下方输入您的API密钥

注意：个人开发者每日有免费调用额度
必须配置API密钥才能使用本工具的功能"""
        
        info_label = tk.Label(dialog, text=info_text, justify=tk.LEFT, wraplength=450)
        info_label.pack(pady=10, padx=10)
        
        # API密钥输入
        key_frame = tk.Frame(dialog)
        key_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(key_frame, text="API密钥:").pack(anchor=tk.W)
        key_entry = tk.Entry(key_frame, width=50, show="*")
        key_entry.pack(fill=tk.X, pady=5)
        
        # 按钮
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=10)
        
        def save_key():
            api_key = key_entry.get().strip()
            if api_key:
                self.config.set_amap_api_key(api_key)
                amap_api.set_api_key(api_key)
                messagebox.showinfo("成功", "API密钥已保存")
                dialog.destroy()
            else:
                messagebox.showerror("错误", "请输入有效的API密钥")
        
        def cancel():
            dialog.destroy()
            # 如果没有API密钥，关闭整个工具
            if not self.config.get_amap_api_key():
                messagebox.showwarning("警告", "未配置API密钥，无法使用高德地图功能")
                self.parent.quit()
        
        tk.Button(button_frame, text="保存", command=save_key).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="取消", command=cancel).pack(side=tk.LEFT, padx=5)
        
        key_entry.focus()
        
    def setup_ui(self):
        # 主框架
        main_frame = tk.Frame(self.parent, bg=self.theme.bg_color)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 标题
        title_label = tk.Label(main_frame, text="高德地图工具", 
                              font=("微软雅黑", 16, "bold"),
                              bg=self.theme.bg_color, fg=self.theme.text_color)
        title_label.pack(pady=(0, 10))
        
        # 坐标系说明
        coord_info_label = tk.Label(main_frame, 
                                   text="注意：请输入WGS-84坐标系的经纬度，程序会自动转换为高德地图使用的GCJ-02坐标系", 
                                   font=("微软雅黑", 10),
                                   bg=self.theme.bg_color, fg=self.theme.accent_color,
                                   wraplength=600)
        coord_info_label.pack(pady=(0, 20))
        
        # 创建选项卡
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # 路径规划选项卡
        self.create_route_planning_tab()
        
        # 批量行政区域查询选项卡
        self.create_batch_geocoding_tab()
        
        # 天气预报选项卡
        self.create_weather_tab()
        
        # 状态栏
        self.create_status_bar(main_frame)
        
    def create_route_planning_tab(self):
        """创建路径规划选项卡"""
        route_frame = ttk.Frame(self.notebook)
        self.notebook.add(route_frame, text="路径规划")
        
        # 起点输入
        origin_frame = tk.Frame(route_frame)
        origin_frame.pack(fill=tk.X, pady=5)
        
        origin_label_frame = tk.Frame(origin_frame)
        origin_label_frame.pack(fill=tk.X)
        
        tk.Label(origin_label_frame, text="起点经纬度 (WGS-84):", 
                font=("微软雅黑", 10), bg=self.theme.bg_color, fg=self.theme.text_color).pack(side=tk.LEFT)
        
        # 收藏位置按钮
        tk.Button(origin_label_frame, text="收藏位置", 
                 command=lambda: self.show_favorite_locations('origin'),
                 font=("微软雅黑", 8)).pack(side=tk.RIGHT, padx=5)
        
        self.origin_entry = tk.Entry(origin_frame, font=("微软雅黑", 10))
        self.origin_entry.pack(fill=tk.X, pady=2)
        self.origin_entry.insert(0, "116.397428,39.90923")  # 默认值：天安门
        
        # 终点输入
        destination_frame = tk.Frame(route_frame)
        destination_frame.pack(fill=tk.X, pady=5)
        
        destination_label_frame = tk.Frame(destination_frame)
        destination_label_frame.pack(fill=tk.X)
        
        tk.Label(destination_label_frame, text="终点经纬度 (WGS-84):", 
                font=("微软雅黑", 10), bg=self.theme.bg_color, fg=self.theme.text_color).pack(side=tk.LEFT)
        
        # 收藏位置按钮
        tk.Button(destination_label_frame, text="收藏位置", 
                 command=lambda: self.show_favorite_locations('destination'),
                 font=("微软雅黑", 8)).pack(side=tk.RIGHT, padx=5)
        
        self.destination_entry = tk.Entry(destination_frame, font=("微软雅黑", 10))
        self.destination_entry.pack(fill=tk.X, pady=2)
        self.destination_entry.insert(0, "116.407526,39.90403")  # 默认值：王府井
        
        # 计算按钮和功能按钮
        btn_frame = tk.Frame(route_frame, bg=self.theme.bg_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 主要功能按钮
        main_button_frame = tk.Frame(btn_frame, bg=self.theme.bg_color)
        main_button_frame.pack()
        
        calculate_btn = tk.Button(main_button_frame, text="计算路径", 
                                 command=self.calculate_routes,
                                 bg=self.theme.button_color, fg="white",
                                 font=("微软雅黑", 10), relief=tk.FLAT)
        calculate_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 辅助功能按钮
        aux_button_frame = tk.Frame(btn_frame, bg=self.theme.bg_color)
        aux_button_frame.pack(pady=5)
        
        tk.Button(aux_button_frame, text="历史记录", 
                 command=self.show_history,
                 font=("微软雅黑", 9)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(aux_button_frame, text="管理收藏", 
                 command=self.manage_favorites,
                 font=("微软雅黑", 9)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(aux_button_frame, text="设置", 
                 command=self.show_settings,
                 font=("微软雅黑", 9)).pack(side=tk.LEFT, padx=5)
        
        # 结果显示区域
        result_frame = tk.Frame(route_frame, bg=self.theme.bg_color)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tk.Label(result_frame, text="路径规划结果:", 
                font=("微软雅黑", 10, "bold"), bg=self.theme.bg_color, fg=self.theme.text_color).pack(anchor=tk.W)
        
        # 创建文本框和滚动条
        text_frame = tk.Frame(result_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.route_result_text = tk.Text(text_frame, height=15, wrap=tk.WORD)
        scrollbar = tk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.route_result_text.yview)
        self.route_result_text.configure(yscrollcommand=scrollbar.set)
        
        self.route_result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def create_batch_geocoding_tab(self):
        """创建批量行政区域查询选项卡"""
        geocoding_frame = ttk.Frame(self.notebook)
        self.notebook.add(geocoding_frame, text="批量行政区域查询")
        
        # 说明文字
        info_label = tk.Label(geocoding_frame, 
                             text="上传Excel文件进行批量行政区域查询\n文件格式：A列为经度，B列为纬度",
                             font=("微软雅黑", 10), bg=self.theme.bg_color, fg=self.theme.text_color,
                             justify=tk.LEFT)
        info_label.pack(pady=10, padx=10, anchor=tk.W)
        
        # 按钮框架
        btn_frame = tk.Frame(geocoding_frame, bg=self.theme.bg_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 下载模板按钮
        download_template_btn = tk.Button(btn_frame, text="下载Excel模板", 
                                         command=self.download_template,
                                         bg=self.theme.button_color, fg="white",
                                         font=("微软雅黑", 10), relief=tk.FLAT)
        download_template_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 上传文件按钮
        upload_btn = tk.Button(btn_frame, text="上传Excel文件", 
                              command=self.upload_excel,
                              bg=self.theme.button_color, fg="white",
                              font=("微软雅黑", 10), relief=tk.FLAT)
        upload_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 开始查询按钮
        query_btn = tk.Button(btn_frame, text="开始查询", 
                             command=self.start_batch_geocoding,
                             bg=self.theme.button_color, fg="white",
                             font=("微软雅黑", 10), relief=tk.FLAT)
        query_btn.pack(side=tk.LEFT)
        
        # 文件路径显示
        self.file_path_var = tk.StringVar(value="未选择文件")
        file_label = tk.Label(geocoding_frame, textvariable=self.file_path_var,
                             font=("微软雅黑", 9), bg=self.theme.bg_color, fg=self.theme.text_color)
        file_label.pack(pady=5, padx=10, anchor=tk.W)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(geocoding_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, padx=10, pady=5)
        
        # 结果显示
        result_frame = tk.Frame(geocoding_frame, bg=self.theme.bg_color)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tk.Label(result_frame, text="查询结果:", 
                font=("微软雅黑", 10, "bold"), bg=self.theme.bg_color, fg=self.theme.text_color).pack(anchor=tk.W)
        
        text_frame = tk.Frame(result_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.geocoding_result_text = tk.Text(text_frame, height=10, wrap=tk.WORD)
        scrollbar2 = tk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.geocoding_result_text.yview)
        self.geocoding_result_text.configure(yscrollcommand=scrollbar2.set)
        
        self.geocoding_result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar2.pack(side=tk.RIGHT, fill=tk.Y)
        
    def create_weather_tab(self):
        """创建天气预报选项卡"""
        weather_frame = ttk.Frame(self.notebook)
        self.notebook.add(weather_frame, text="天气预报")
        
        # 城市输入
        input_frame = tk.Frame(weather_frame, bg=self.theme.bg_color)
        input_frame.pack(fill=tk.X, padx=10, pady=10)
        
        tk.Label(input_frame, text="城市名称:", 
                font=("微软雅黑", 10), bg=self.theme.bg_color, fg=self.theme.text_color).pack(side=tk.LEFT)
        
        self.city_entry = tk.Entry(input_frame, width=20)
        self.city_entry.pack(side=tk.LEFT, padx=(10, 10))
        self.city_entry.insert(0, "北京")
        
        query_weather_btn = tk.Button(input_frame, text="查询天气", 
                                     command=self.query_weather,
                                     bg=self.theme.button_color, fg="white",
                                     font=("微软雅黑", 10), relief=tk.FLAT)
        query_weather_btn.pack(side=tk.LEFT)
        
        # 天气结果显示
        result_frame = tk.Frame(weather_frame, bg=self.theme.bg_color)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tk.Label(result_frame, text="天气信息:", 
                font=("微软雅黑", 10, "bold"), bg=self.theme.bg_color, fg=self.theme.text_color).pack(anchor=tk.W)
        
        text_frame = tk.Frame(result_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.weather_result_text = tk.Text(text_frame, height=15, wrap=tk.WORD)
        scrollbar3 = tk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.weather_result_text.yview)
        self.weather_result_text.configure(yscrollcommand=scrollbar3.set)
        
        self.weather_result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar3.pack(side=tk.RIGHT, fill=tk.Y)
        
    def create_status_bar(self, parent):
        """创建状态栏"""
        status_frame = tk.Frame(parent, bg=self.theme.bg_color, relief=tk.SUNKEN, bd=1)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(10, 0))
        
        status_label = tk.Label(status_frame, textvariable=self.status_text,
                               font=("微软雅黑", 9), bg=self.theme.bg_color, fg=self.theme.text_color,
                               anchor=tk.W)
        status_label.pack(fill=tk.X, padx=5, pady=2)
        
    def update_status(self, message):
        """更新状态栏"""
        self.status_text.set(message)
        self.parent.update_idletasks()
        
    def calculate_routes(self):
        """计算路径规划"""
        try:
            # 从输入框获取坐标
            origin_coords = self.origin_entry.get().strip()
            destination_coords = self.destination_entry.get().strip()
            
            if not origin_coords or not destination_coords:
                messagebox.showerror("错误", "请输入起点和终点坐标")
                return
                
            # 验证坐标格式
            try:
                start_lng, start_lat = map(float, origin_coords.split(','))
                end_lng, end_lat = map(float, destination_coords.split(','))
            except ValueError:
                messagebox.showerror("错误", "坐标格式错误，请使用：经度,纬度")
                return
            
            self.update_status("正在计算路径...")
            
            # 在新线程中执行路径计算
            thread = threading.Thread(target=self._calculate_routes_thread, 
                                    args=(start_lng, start_lat, end_lng, end_lat))
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            messagebox.showerror("错误", f"计算路径时出错: {str(e)}")
            self.update_status("就绪")
            
    def _calculate_routes_thread(self, start_lng, start_lat, end_lng, end_lat):
        """在线程中计算路径"""
        try:
            # 将WGS-84坐标转换为GCJ-02坐标（高德地图使用的坐标系）
            start_lng_gcj, start_lat_gcj = wgs84_to_gcj02(start_lng, start_lat)
            end_lng_gcj, end_lat_gcj = wgs84_to_gcj02(end_lng, end_lat)
            
            origin = f"{start_lng_gcj},{start_lat_gcj}"
            destination = f"{end_lng_gcj},{end_lat_gcj}"
            
            # 在结果中显示坐标转换信息
            coord_info = f"坐标转换信息:\n起点: WGS-84({start_lng:.6f}, {start_lat:.6f}) -> GCJ-02({start_lng_gcj:.6f}, {start_lat_gcj:.6f})\n终点: WGS-84({end_lng:.6f}, {end_lat:.6f}) -> GCJ-02({end_lng_gcj:.6f}, {end_lat_gcj:.6f})\n\n"
            
            results = [coord_info]
            
            # 首先获取起点和终点的地址信息
            self.parent.after(0, lambda: self.update_status("正在获取起点地址..."))
            start_address_result = self._call_amap_api("maps_regeocode", {
                "location": origin
            })
            
            self.parent.after(0, lambda: self.update_status("正在获取终点地址..."))
            end_address_result = self._call_amap_api("maps_regeocode", {
                "location": destination
            })
            
            # 显示地址信息
            if start_address_result:
                results.append(f"起点地址信息:\n{start_address_result}\n")
            if end_address_result:
                results.append(f"终点地址信息:\n{end_address_result}\n")
            
            # 使用MCP服务器调用高德API
            # 1. 首先计算直线距离
            self.parent.after(0, lambda: self.update_status("正在计算直线距离..."))
            straight_distance_result = self._call_amap_api("maps_distance", {
                "origins": origin,
                "destination": destination,
                "type": "0"  # 直线距离
            })
            if straight_distance_result:
                results.append(f"直线距离:\n{straight_distance_result}\n")
            
            # 2. 然后计算驾车路径和距离
            self.parent.after(0, lambda: self.update_status("正在计算驾车路径..."))
            driving_result = self._call_amap_api("maps_direction_driving", {
                "origin": origin,
                "destination": destination
            })
            if driving_result:
                results.append(f"驾车路径:\n{driving_result}\n")
            
            # 3. 计算驾车距离
            self.parent.after(0, lambda: self.update_status("正在计算驾车距离..."))
            driving_distance_result = self._call_amap_api("maps_distance", {
                "origins": origin,
                "destination": destination,
                "type": "1"  # 驾车距离
            })
            if driving_distance_result:
                results.append(f"驾车距离:\n{driving_distance_result}\n")
            
            # 保存到历史记录
            try:
                self.history_manager.add_record(
                    "route_planning",
                    {
                        "origin": origin,
                        "destination": destination,
                        "origin_address": start_address_result.split('\n')[0] if start_address_result else "未知地址",
                        "destination_address": end_address_result.split('\n')[0] if end_address_result else "未知地址"
                    }
                )
            except Exception as history_error:
                print(f"保存历史记录失败: {history_error}")
            
            # 更新UI
            result_text = "\n".join(results) if results else "路径计算失败，请检查网络连接和坐标有效性"
            self.parent.after(0, lambda: self._update_route_result(result_text))
            self.parent.after(0, lambda: self.update_status("路径计算完成"))
            
        except Exception as e:
            error_msg = f"路径计算出错: {str(e)}"
            self.parent.after(0, lambda: self._update_route_result(error_msg))
            self.parent.after(0, lambda: self.update_status("计算失败"))
            
    def _update_route_result(self, text):
        """更新路径结果显示"""
        self.route_result_text.delete(1.0, tk.END)
        self.route_result_text.insert(tk.END, text)
        
    def _call_amap_api(self, tool_name, params):
        """调用高德API（仅使用真实API）"""
        try:
            # 检查是否配置了API密钥
            api_key = self.config.get_amap_api_key()
            
            if not api_key:
                return "错误：未配置API密钥，请先配置高德地图API密钥"
            
            # 调用真实API
            if tool_name == "maps_direction_driving":
                result = amap_api.direction_driving(params['origin'], params['destination'])
                if result['status'] == 'success':
                    return self._format_real_driving_result(result)
                else:
                    return f"驾车路径查询失败: {result.get('message', '未知错误')}"
            elif tool_name == "maps_direction_walking":
                result = amap_api.direction_walking(params['origin'], params['destination'])
                if result['status'] == 'success':
                    return self._format_real_walking_result(result)
                else:
                    return f"步行路径查询失败: {result.get('message', '未知错误')}"
            elif tool_name == "maps_distance":
                distance_type = int(params.get('type', '1'))
                result = amap_api.distance(params['origins'], params['destination'], distance_type)
                if result['status'] == 'success':
                    return self._format_real_distance_result(result, distance_type)
                else:
                    return f"距离查询失败: {result.get('message', '未知错误')}"
            elif tool_name == "maps_regeocode":
                lng, lat = map(float, params['location'].split(','))
                result = amap_api.regeocode(lng, lat)
                if result['status'] == 'success':
                    return self._format_real_regeocode_result(result)
                else:
                    return f"逆地理编码查询失败: {result.get('message', '未知错误')}"
            elif tool_name == "maps_weather":
                result = amap_api.weather(params['city'])
                if result['status'] == 'success':
                    return self._format_real_weather_result(result)
                else:
                    return f"天气查询失败: {result.get('message', '未知错误')}"
            else:
                return f"不支持的API工具: {tool_name}"
                
        except Exception as e:
            return f"API调用失败: {str(e)}"
    
    # 真实API结果格式化方法
    def _format_real_driving_result(self, result):
        """格式化真实驾车路径结果"""
        try:
            distance = float(result['distance']) / 1000  # 转换为公里
            duration = int(result['duration']) // 60  # 转换为分钟
            
            steps_text = []
            for step in result.get('steps', []):
                instruction = step['instruction']
                step_distance = float(step['distance'])
                if step_distance >= 1000:
                    distance_str = f"{step_distance/1000:.1f}公里"
                else:
                    distance_str = f"{step_distance:.0f}米"
                steps_text.append(f"• {instruction} ({distance_str})")
            
            return f"""驾车路径规划：
距离：{distance:.1f}公里
预计用时：{duration}分钟

详细路线：
{chr(10).join(steps_text)}"""
        except Exception as e:
            return f"驾车路径解析失败: {str(e)}"
    
    def _format_real_walking_result(self, result):
        """格式化真实步行路径结果"""
        try:
            distance = float(result['distance'])  # 米
            duration = int(result['duration']) // 60  # 转换为分钟
            
            if distance >= 1000:
                distance_str = f"{distance/1000:.1f}公里"
            else:
                distance_str = f"{distance:.0f}米"
            
            steps_text = []
            for step in result.get('steps', []):
                instruction = step['instruction']
                step_distance = float(step['distance'])
                if step_distance >= 1000:
                    step_distance_str = f"{step_distance/1000:.1f}公里"
                else:
                    step_distance_str = f"{step_distance:.0f}米"
                steps_text.append(f"• {instruction} ({step_distance_str})")
            
            return f"""步行路径规划：
距离：{distance_str}
预计用时：{duration}分钟

详细路线：
{chr(10).join(steps_text)}"""
        except Exception as e:
            return f"步行路径解析失败: {str(e)}"
    
    def _format_real_distance_result(self, result, distance_type):
        """格式化真实距离测量结果"""
        try:
            distance = float(result['distance'])
            
            if distance >= 1000:
                distance_str = f"{distance/1000:.1f}公里"
            else:
                distance_str = f"{distance:.0f}米"
            
            type_names = {0: "直线距离", 1: "驾车距离", 3: "步行距离"}
            type_name = type_names.get(distance_type, "距离")
            
            if distance_type == 1 and 'duration' in result:
                duration = int(result['duration']) // 60
                return f"{type_name}：{distance_str}\n预计用时：{duration}分钟"
            else:
                return f"{type_name}：{distance_str}"
        except Exception as e:
            return f"距离测量解析失败: {str(e)}"
    
    def _format_real_regeocode_result(self, result):
        """格式化真实逆地理编码结果"""
        try:
            province = result.get('province', '')
            city = result.get('city', '')
            district = result.get('district', '')
            township = result.get('township', '')
            
            formatted_address = result.get('formatted_address', '')
            
            # 获取POI信息
            pois = result.get('pois', [])
            poi_info = ""
            if pois:
                nearest_poi = pois[0]
                poi_name = nearest_poi.get('name', '')
                poi_distance = nearest_poi.get('distance', '')
                if poi_name and poi_distance:
                    poi_info = f"\n附近地标：{poi_name} (距离{poi_distance}米)"
            
            return f"""地址信息：
省份：{province}
城市：{city}
区县：{district}
街道：{township}
详细地址：{formatted_address}{poi_info}"""
        except Exception as e:
            return f"地址解析失败: {str(e)}"
    
    def _format_real_weather_result(self, result):
        """格式化真实天气查询结果"""
        try:
            city = result.get('city', '')
            province = result.get('province', '')
            reporttime = result.get('reporttime', '')
            
            # 获取天气预报信息
            casts = result.get('casts', [])
            if casts:
                today_weather = casts[0]
                date = today_weather.get('date', '')
                week = today_weather.get('week', '')
                dayweather = today_weather.get('dayweather', '')
                nightweather = today_weather.get('nightweather', '')
                daytemp = today_weather.get('daytemp', '')
                nighttemp = today_weather.get('nighttemp', '')
                daywind = today_weather.get('daywind', '')
                nightwind = today_weather.get('nightwind', '')
                daypower = today_weather.get('daypower', '')
                nightpower = today_weather.get('nightpower', '')
                
                return f"""天气信息：
城市：{province} {city}
日期：{date} {week}
白天：{dayweather} {daytemp}°C {daywind}{daypower}
夜间：{nightweather} {nighttemp}°C {nightwind}{nightpower}
更新时间：{reporttime}"""
            else:
                return f"""天气信息：
城市：{province} {city}
更新时间：{reporttime}
暂无详细天气数据"""
        except Exception as e:
            return f"天气信息解析失败: {str(e)}"
    
    # 新增功能方法
    def show_favorite_locations(self, entry_type):
        """显示收藏位置选择对话框"""
        try:
            # 根据entry_type确定对应的Entry对象
            if entry_type == 'origin':
                entry_widget = self.origin_entry
            elif entry_type == 'destination':
                entry_widget = self.destination_entry
            else:
                return  # 无效的entry_type
            
            # 创建选择对话框
            dialog = tk.Toplevel(self.parent)
            dialog.title("选择位置")
            dialog.geometry("400x300")
            dialog.transient(self.parent)
            dialog.grab_set()
            
            # 居中显示
            dialog.geometry("+%d+%d" % (self.parent.winfo_rootx() + 50, self.parent.winfo_rooty() + 50))
            
            # 创建标签页
            notebook = ttk.Notebook(dialog)
            notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 收藏位置标签页
            favorites_frame = ttk.Frame(notebook)
            notebook.add(favorites_frame, text="收藏位置")
            
            # 收藏位置列表
            favorites_listbox = tk.Listbox(favorites_frame)
            favorites_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # 加载收藏位置
            favorite_locations = self.config.get('favorite_locations', [])
            for location in favorite_locations:
                if isinstance(location, dict):
                    name = location.get('name', '未命名')
                    lng = location.get('lng', 0)
                    lat = location.get('lat', 0)
                    coords = f"{lng},{lat}"
                    favorites_listbox.insert(tk.END, f"{name} ({coords})")
            
            # 历史记录标签页
            history_frame = ttk.Frame(notebook)
            notebook.add(history_frame, text="历史记录")
            
            # 历史记录列表
            history_listbox = tk.Listbox(history_frame)
            history_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # 加载历史记录
            recent_locations = self.history_manager.get_recent_locations(20)
            for location in recent_locations:
                coords = location.get('coordinates', '')
                address = location.get('address', '未知地址')
                history_listbox.insert(tk.END, f"{address} ({coords})")
            
            # 按钮框架
            button_frame = tk.Frame(dialog)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            def on_select():
                try:
                    current_tab = notebook.index(notebook.select())
                    if current_tab == 0:  # 收藏位置
                        selection = favorites_listbox.curselection()
                        if selection:
                            text = favorites_listbox.get(selection[0])
                            # 提取坐标：格式为 "名称 (经度,纬度)"
                            if '(' in text and ')' in text:
                                coords = text.split('(')[1].split(')')[0]
                                entry_widget.delete(0, tk.END)
                                entry_widget.insert(0, coords)
                                dialog.destroy()
                            else:
                                messagebox.showwarning("警告", "无法解析坐标格式")
                    else:  # 历史记录
                        selection = history_listbox.curselection()
                        if selection:
                            text = history_listbox.get(selection[0])
                            # 提取坐标：格式为 "地址 (经度,纬度)"
                            if '(' in text and ')' in text:
                                coords = text.split('(')[1].split(')')[0]
                                entry_widget.delete(0, tk.END)
                                entry_widget.insert(0, coords)
                                dialog.destroy()
                            else:
                                messagebox.showwarning("警告", "无法解析坐标格式")
                except Exception as e:
                    messagebox.showerror("错误", f"选择位置时出错: {str(e)}")
            
            tk.Button(button_frame, text="选择", command=on_select).pack(side=tk.LEFT, padx=5)
            tk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("错误", f"显示位置选择失败: {str(e)}")
    
    def show_history(self):
        """显示历史记录管理窗口"""
        try:
            # 创建历史记录窗口
            history_window = tk.Toplevel(self.parent)
            history_window.title("历史记录")
            history_window.geometry("600x400")
            history_window.transient(self.parent)
            
            # 居中显示
            history_window.geometry("+%d+%d" % (self.parent.winfo_rootx() + 50, self.parent.winfo_rooty() + 50))
            
            # 创建列表框
            listbox_frame = tk.Frame(history_window)
            listbox_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 添加滚动条
            scrollbar = tk.Scrollbar(listbox_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            history_listbox = tk.Listbox(listbox_frame, yscrollcommand=scrollbar.set)
            history_listbox.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=history_listbox.yview)
            
            # 加载历史记录
            def load_history():
                history_listbox.delete(0, tk.END)
                records = self.history_manager.get_all_records()
                for record in records:
                    timestamp = record.get('timestamp', '')
                    query_type = record.get('query_type', '')
                    origin = record.get('origin', '')
                    destination = record.get('destination', '')
                    
                    if query_type == 'route_planning':
                        display_text = f"[{timestamp}] 路径规划: {origin} → {destination}"
                    else:
                        display_text = f"[{timestamp}] {query_type}: {origin}"
                    
                    history_listbox.insert(tk.END, display_text)
            
            load_history()
            
            # 按钮框架
            button_frame = tk.Frame(history_window)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            def delete_selected():
                selection = history_listbox.curselection()
                if selection:
                    if messagebox.askyesno("确认删除", "确定要删除选中的历史记录吗？"):
                        # 这里需要实现删除逻辑
                        load_history()
            
            def clear_all():
                if messagebox.askyesno("确认清空", "确定要清空所有历史记录吗？"):
                    self.history_manager.clear_all_records()
                    load_history()
            
            tk.Button(button_frame, text="删除选中", command=delete_selected).pack(side=tk.LEFT, padx=5)
            tk.Button(button_frame, text="清空全部", command=clear_all).pack(side=tk.LEFT, padx=5)
            tk.Button(button_frame, text="关闭", command=history_window.destroy).pack(side=tk.RIGHT, padx=5)
            
        except Exception as e:
            messagebox.showerror("错误", f"显示历史记录失败: {str(e)}")
    
    def manage_favorites(self):
        """管理收藏位置"""
        try:
            # 创建收藏管理窗口
            favorites_window = tk.Toplevel(self.parent)
            favorites_window.title("管理收藏位置")
            favorites_window.geometry("500x400")
            favorites_window.transient(self.parent)
            
            # 居中显示
            favorites_window.geometry("+%d+%d" % (self.parent.winfo_rootx() + 50, self.parent.winfo_rooty() + 50))
            
            # 收藏列表
            list_frame = tk.Frame(favorites_window)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 添加滚动条
            scrollbar = tk.Scrollbar(list_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            favorites_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
            favorites_listbox.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=favorites_listbox.yview)
            
            # 加载收藏位置
            def load_favorites():
                favorites_listbox.delete(0, tk.END)
                favorite_locations = self.config.get('favorite_locations', [])
                for location in favorite_locations:
                    if isinstance(location, dict):
                        name = location.get('name', '未命名')
                        lng = location.get('lng', 0)
                        lat = location.get('lat', 0)
                        coords = f"{lng},{lat}"
                        favorites_listbox.insert(tk.END, f"{name}: {coords}")
            
            load_favorites()
            
            # 添加新收藏
            add_frame = tk.Frame(favorites_window)
            add_frame.pack(fill=tk.X, padx=10, pady=5)
            
            tk.Label(add_frame, text="名称:").pack(side=tk.LEFT)
            name_entry = tk.Entry(add_frame, width=15)
            name_entry.pack(side=tk.LEFT, padx=5)
            
            tk.Label(add_frame, text="坐标:").pack(side=tk.LEFT)
            coords_entry = tk.Entry(add_frame, width=20)
            coords_entry.pack(side=tk.LEFT, padx=5)
            
            def add_favorite():
                name = name_entry.get().strip()
                coords = coords_entry.get().strip()
                if name and coords:
                    try:
                        # 解析坐标
                        lng, lat = map(float, coords.split(','))
                        
                        favorite_locations = self.config.get('favorite_locations', [])
                        # 检查是否已存在同名位置
                        for i, location in enumerate(favorite_locations):
                            if location.get('name') == name:
                                favorite_locations[i] = {'name': name, 'lng': lng, 'lat': lat}
                                break
                        else:
                            favorite_locations.append({'name': name, 'lng': lng, 'lat': lat})
                        
                        self.config.set('favorite_locations', favorite_locations)
                        self.config.save_config()
                        name_entry.delete(0, tk.END)
                        coords_entry.delete(0, tk.END)
                        load_favorites()
                    except ValueError:
                        messagebox.showerror("错误", "坐标格式错误，请使用 经度,纬度 格式")
                else:
                    messagebox.showwarning("警告", "请输入名称和坐标")
            
            tk.Button(add_frame, text="添加", command=add_favorite).pack(side=tk.LEFT, padx=5)
            
            # 按钮框架
            button_frame = tk.Frame(favorites_window)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            def delete_selected():
                selection = favorites_listbox.curselection()
                if selection:
                    if messagebox.askyesno("确认删除", "确定要删除选中的收藏位置吗？"):
                        text = favorites_listbox.get(selection[0])
                        name = text.split(':')[0]
                        favorite_locations = self.config.get('favorite_locations', [])
                        # 找到并删除对应的位置
                        for i, location in enumerate(favorite_locations):
                            if location.get('name') == name:
                                favorite_locations.pop(i)
                                break
                        
                        self.config.set('favorite_locations', favorite_locations)
                        self.config.save_config()
                        load_favorites()
            
            tk.Button(button_frame, text="删除选中", command=delete_selected).pack(side=tk.LEFT, padx=5)
            tk.Button(button_frame, text="关闭", command=favorites_window.destroy).pack(side=tk.RIGHT, padx=5)
            
        except Exception as e:
            messagebox.showerror("错误", f"管理收藏位置失败: {str(e)}")
    
    def show_settings(self):
        """显示设置窗口"""
        try:
            # 创建设置窗口
            settings_window = tk.Toplevel(self.parent)
            settings_window.title("设置")
            settings_window.geometry("400x300")
            settings_window.transient(self.parent)
            
            # 居中显示
            settings_window.geometry("+%d+%d" % (self.parent.winfo_rootx() + 50, self.parent.winfo_rooty() + 50))
            
            # API密钥设置
            api_frame = tk.LabelFrame(settings_window, text="API设置", padx=10, pady=10)
            api_frame.pack(fill=tk.X, padx=10, pady=10)
            
            tk.Label(api_frame, text="高德地图API密钥:").pack(anchor=tk.W)
            api_key_entry = tk.Entry(api_frame, width=50, show="*")
            api_key_entry.pack(fill=tk.X, pady=5)
            
            # 加载当前API密钥
            current_key = self.config.get_amap_api_key()
            if current_key:
                api_key_entry.insert(0, current_key)
            
            # 缓存设置
            cache_frame = tk.LabelFrame(settings_window, text="缓存设置", padx=10, pady=10)
            cache_frame.pack(fill=tk.X, padx=10, pady=10)
            
            cache_enabled_var = tk.BooleanVar()
            cache_enabled_var.set(self.config.get('cache_enabled', True))
            tk.Checkbutton(cache_frame, text="启用结果缓存", variable=cache_enabled_var).pack(anchor=tk.W)
            
            tk.Label(cache_frame, text="缓存过期时间(小时):").pack(anchor=tk.W)
            cache_ttl_entry = tk.Entry(cache_frame, width=10)
            cache_ttl_entry.pack(anchor=tk.W, pady=5)
            cache_ttl_entry.insert(0, str(self.config.get('cache_ttl_hours', 24)))
            
            # 历史记录设置
            history_frame = tk.LabelFrame(settings_window, text="历史记录设置", padx=10, pady=10)
            history_frame.pack(fill=tk.X, padx=10, pady=10)
            
            tk.Label(history_frame, text="最大历史记录数:").pack(anchor=tk.W)
            max_history_entry = tk.Entry(history_frame, width=10)
            max_history_entry.pack(anchor=tk.W, pady=5)
            max_history_entry.insert(0, str(self.config.get('max_history_records', 1000)))
            
            # 按钮框架
            button_frame = tk.Frame(settings_window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            def save_settings():
                try:
                    # 保存API密钥
                    new_api_key = api_key_entry.get().strip()
                    if new_api_key:
                        self.config.set_amap_api_key(new_api_key)
                        # 重新设置API密钥
                        amap_api.set_api_key(new_api_key)
                    
                    # 保存其他设置
                    self.config.set('cache_enabled', cache_enabled_var.get())
                    self.config.set('cache_ttl_hours', int(cache_ttl_entry.get()))
                    self.config.set('max_history_records', int(max_history_entry.get()))
                    
                    self.config.save_config()
                    messagebox.showinfo("成功", "设置已保存")
                    settings_window.destroy()
                except Exception as e:
                    messagebox.showerror("错误", f"保存设置失败: {str(e)}")
            
            tk.Button(button_frame, text="保存", command=save_settings).pack(side=tk.LEFT, padx=5)
            tk.Button(button_frame, text="取消", command=settings_window.destroy).pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("错误", f"显示设置失败: {str(e)}")
            









            
    def download_template(self):
        """下载Excel模板"""
        try:
            # 创建Excel模板
            wb = Workbook()
            ws = wb.active
            ws.title = "坐标模板"
            
            # 设置表头
            ws['A1'] = "经度(WGS-84)"
            ws['B1'] = "纬度(WGS-84)"
            ws['C1'] = "省份"
            ws['D1'] = "城市"
            ws['E1'] = "区县"
            ws['F1'] = "详细地址"
            
            # 添加坐标系说明
            ws['A3'] = "说明：请输入WGS-84坐标系的经纬度"
            ws['B3'] = "程序会自动转换为高德地图的GCJ-02坐标系"
            ws['A4'] = "示例：116.397428（天安门经度）"
            ws['B4'] = "示例：39.90923（天安门纬度）"
            
            # 添加示例数据（仅作为格式参考）
            ws['A5'] = 116.397428
            ws['B5'] = 39.90923
            ws['C5'] = "（查询后自动填充）"
            ws['D5'] = "（查询后自动填充）"
            ws['E5'] = "（查询后自动填充）"
            ws['F5'] = "（查询后自动填充）"
            
            # 保存文件
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                title="保存Excel模板"
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("成功", f"模板已保存到: {file_path}")
                self.update_status("模板下载完成")
            
        except Exception as e:
            messagebox.showerror("错误", f"模板下载失败: {str(e)}")
            self.update_status("模板下载失败")
            
    def upload_excel(self):
        """上传Excel文件"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel文件", "*.xlsx"), ("Excel文件", "*.xls")],
            title="选择Excel文件"
        )
        
        if file_path:
            self.file_path_var.set(f"已选择: {os.path.basename(file_path)}")
            self.excel_file_path = file_path
            self.update_status("文件上传完成")
        
    def start_batch_geocoding(self):
        """开始批量地理编码"""
        if not hasattr(self, 'excel_file_path'):
            messagebox.showerror("错误", "请先上传Excel文件")
            return
            
        self.update_status("正在进行批量查询...")
        
        # 在新线程中执行批量查询
        thread = threading.Thread(target=self._batch_geocoding_thread)
        thread.daemon = True
        thread.start()
        
    def _batch_geocoding_thread(self):
        """批量地理编码线程"""
        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(self.excel_file_path)
            ws = wb.active
            
            # 获取数据行数
            max_row = ws.max_row
            results = []
            
            for row in range(6, max_row + 1):  # 从第6行开始（跳过表头和说明行）
                lng = ws.cell(row=row, column=1).value
                lat = ws.cell(row=row, column=2).value
                
                if lng is not None and lat is not None:
                    try:
                        # 将WGS-84坐标转换为GCJ-02坐标（高德地图使用的坐标系）
                        lng_gcj, lat_gcj = wgs84_to_gcj02(lng, lat)
                        location = f"{lng_gcj},{lat_gcj}"
                        
                        # 调用逆地理编码API
                        result = self._call_amap_api("maps_regeocode", {
                            "location": location
                        })
                        
                        # 在结果中记录坐标转换信息
                        coord_conversion = f"坐标转换: WGS-84({lng:.6f}, {lat:.6f}) -> GCJ-02({lng_gcj:.6f}, {lat_gcj:.6f})"
                        
                        # 解析API返回结果并更新Excel
                        if "错误" not in result and "失败" not in result:
                            # 尝试从结果中提取地址信息
                            lines = result.split('\n')
                            province = city = district = address = ""
                            
                            for line in lines:
                                if line.startswith('省份：'):
                                    province = line.replace('省份：', '').strip()
                                elif line.startswith('城市：'):
                                    city = line.replace('城市：', '').strip()
                                elif line.startswith('区县：'):
                                    district = line.replace('区县：', '').strip()
                                elif line.startswith('详细地址：'):
                                    address = line.replace('详细地址：', '').strip()
                            
                            ws.cell(row=row, column=3, value=province)
                            ws.cell(row=row, column=4, value=city)
                            ws.cell(row=row, column=5, value=district)
                            ws.cell(row=row, column=6, value=address)
                        else:
                            # API调用失败，记录错误信息
                            ws.cell(row=row, column=3, value="查询失败")
                            ws.cell(row=row, column=4, value="查询失败")
                            ws.cell(row=row, column=5, value="查询失败")
                            ws.cell(row=row, column=6, value=result)
                        
                        results.append(f"行{row}: {coord_conversion} -> {result}")
                        
                    except Exception as e:
                        results.append(f"行{row}: 查询失败 - {str(e)}")
                
                # 更新进度
                progress = ((row - 1) / (max_row - 1)) * 100
                self.parent.after(0, lambda p=progress: self.progress_var.set(p))
                
            # 保存结果文件
            output_path = self.excel_file_path.replace('.xlsx', '_结果.xlsx')
            wb.save(output_path)
            
            # 更新UI
            result_text = "\n".join(results)
            self.parent.after(0, lambda: self._update_geocoding_result(result_text))
            self.parent.after(0, lambda: messagebox.showinfo("完成", f"查询完成，结果已保存到: {output_path}"))
            self.parent.after(0, lambda: self.update_status("批量查询完成"))
            
        except Exception as e:
            error_msg = f"批量查询出错: {str(e)}"
            self.parent.after(0, lambda: self._update_geocoding_result(error_msg))
            self.parent.after(0, lambda: self.update_status("查询失败"))
            
    def _update_geocoding_result(self, text):
        """更新地理编码结果显示"""
        self.geocoding_result_text.delete(1.0, tk.END)
        self.geocoding_result_text.insert(tk.END, text)
        
    def query_weather(self):
        """查询天气"""
        city = self.city_entry.get().strip()
        if not city:
            messagebox.showerror("错误", "请输入城市名称")
            return
            
        self.update_status("正在查询天气...")
        
        # 在新线程中查询天气
        thread = threading.Thread(target=self._query_weather_thread, args=(city,))
        thread.daemon = True
        thread.start()
        
    def _query_weather_thread(self, city):
        """天气查询线程"""
        try:
            result = self._call_amap_api("maps_weather", {
                "city": city
            })
            
            weather_info = f"城市: {city}\n天气信息: {result}"
            
            self.parent.after(0, lambda: self._update_weather_result(weather_info))
            self.parent.after(0, lambda: self.update_status("天气查询完成"))
            
        except Exception as e:
            error_msg = f"天气查询出错: {str(e)}"
            self.parent.after(0, lambda: self._update_weather_result(error_msg))
            self.parent.after(0, lambda: self.update_status("天气查询失败"))
            
    def _update_weather_result(self, text):
        """更新天气结果显示"""
        self.weather_result_text.delete(1.0, tk.END)
        self.weather_result_text.insert(tk.END, text)
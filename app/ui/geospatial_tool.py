#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests
import json
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import openpyxl # For Excel export
import requests.utils # For URL encoding
import xml.etree.ElementTree as ET # For KML parsing/generation
from xml.dom.minidom import parseString # For pretty KML output
import os
import sys

# 添加项目根目录到路径，以便导入config模块
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from config import config 

# --- 坐标转换函数 (WGS-84 <-> GCJ-02) ---
x_pi = 3.14159265358979324 * 3000.0 / 180.0
pi = 3.1415926535897932384626  # π
a = 6378245.0  # 长半轴
ee = 0.00669342162296594323  # 扁率

def _transform_lat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + \
          0.1 * lng * lat + 0.2 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * pi) + 40.0 *
            math.sin(lat / 3.0 * pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * pi) + 320 *
            math.sin(lat * pi / 30.0)) * 2.0 / 3.0
    return ret

def _transform_lng(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + \
          0.1 * lng * lat + 0.1 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * pi) + 40.0 *
            math.sin(lng / 3.0 * pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * pi) + 300.0 *
            math.sin(lng / 30.0 * pi)) * 2.0 / 3.0
    return ret

def wgs84_to_gcj02(lng_wgs, lat_wgs):
    """WGS84坐标系转GCJ02坐标系（高德、谷歌中国等）"""
    if not (73.66 < lng_wgs < 135.05 and 3.86 < lat_wgs < 53.55):
        # print("坐标超出中国范围，不进行转换")
        return lng_wgs, lat_wgs
    dlat = _transform_lat(lng_wgs - 105.0, lat_wgs - 35.0)
    dlng = _transform_lng(lng_wgs - 105.0, lat_wgs - 35.0)
    radlat = lat_wgs / 180.0 * pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)
    mglat = lat_wgs + dlat
    mglng = lng_wgs + dlng
    return mglng, mglat

def gcj02_to_wgs84(lng_gcj, lat_gcj):
    """GCJ02坐标系转WGS84坐标系"""
    if not (73.66 < lng_gcj < 135.05 and 3.86 < lat_gcj < 53.55):
        # print("坐标超出中国范围，不进行转换")
        return lng_gcj, lat_gcj
    # dlat = _transform_lat(lng_gcj - 105.0, lat_gcj - 35.0)
    # dlng = _transform_lng(lng_gcj - 105.0, lat_gcj - 35.0)
    # radlat = lat_gcj / 180.0 * pi
    # magic = math.sin(radlat)
    # magic = 1 - ee * magic * magic
    # sqrtmagic = math.sqrt(magic)
    # dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
    # dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)
    # lng_wgs = lng_gcj - dlng
    # lat_wgs = lat_gcj - dlat
    # return lng_wgs, lat_wgs
    # Corrected GCJ02 to WGS84 conversion
    # The original implementation was an approximation. A more accurate way is iterative or using the reverse calculation based on wgs84_to_gcj02
    # For simplicity and common practice, many implementations use a delta method:
    mglng, mglat = wgs84_to_gcj02(lng_gcj, lat_gcj) # Estimate WGS84 by assuming input is WGS84 and see what GCJ02 it produces
    lng_wgs = lng_gcj * 2 - mglng
    lat_wgs = lat_gcj * 2 - mglat
    return lng_wgs, lat_wgs

# --- 高德API调用函数 ---
def get_address_from_amap(lon_wgs, lat_wgs, api_key=None):
    """使用高德逆地理编码API获取地址信息 (输入WGS-84, API使用GCJ-02)"""
    if api_key is None:
        api_key = config.get_amap_api_key()
    if not api_key:
        return "错误：请在配置中设置您的高德API Key"
    lon_gcj, lat_gcj = wgs84_to_gcj02(lon_wgs, lat_wgs)
    url = f"https://restapi.amap.com/v3/geocode/regeo?output=json&location={lon_gcj},{lat_gcj}&key={api_key}&radius=1000&extensions=base"
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status() # 如果请求失败则抛出HTTPError
        data = response.json()
        if data.get("status") == "1" and data.get("regeocode"):
            return data["regeocode"].get("formatted_address", "地址未找到")
        else:
            return f"高德API错误: {data.get('info', '未知错误')}"
    except requests.exceptions.RequestException as e:
        return f"网络请求错误: {e}"
    except json.JSONDecodeError:
        return "解析高德API响应失败"

def get_coords_from_amap(address, city, api_key=None):
    """使用高德地理编码API获取经纬度信息 (返回WGS-84)"""
    if api_key is None:
        api_key = config.get_amap_api_key()
    if not api_key:
        return "错误：请在配置中设置您的高德API Key", None, None
    url = f"https://restapi.amap.com/v3/geocode/geo?address={requests.utils.quote(address)}&city={requests.utils.quote(city)}&output=json&key={api_key}"
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get("status") == "1" and data.get("geocodes"):
            location_gcj_str = data["geocodes"][0].get("location")
            if location_gcj_str:
                lon_gcj_str, lat_gcj_str = location_gcj_str.split(',')
                lon_wgs, lat_wgs = gcj02_to_wgs84(float(lon_gcj_str), float(lat_gcj_str))
                return None, lon_wgs, lat_wgs
            else:
                return "未找到经纬度", None, None
        else:
            return f"高德API错误: {data.get('info', '未知错误')}", None, None
    except requests.exceptions.RequestException as e:
        return f"网络请求错误: {e}", None, None
    except (json.JSONDecodeError, IndexError):
        return "解析高德API响应或数据格式错误", None, None


def search_nearby_pois_amap(lon_wgs, lat_wgs, radius_meters, search_keywords, search_types="", api_key=None):
    """使用高德周边搜索API查找POI (输入WGS-84, API使用GCJ-02, 输出WGS-84)"""
    if api_key is None:
        api_key = config.get_amap_api_key()
    if not api_key:
        return "错误：请在配置中设置您的高德API Key", []
    
    lon_gcj, lat_gcj = wgs84_to_gcj02(lon_wgs, lat_wgs)

    url_params = {
        "key": api_key,
        "location": f"{lon_gcj},{lat_gcj}",
        "keywords": search_keywords,
        "radius": str(radius_meters),
        "offset": "25",
        "page": "1",
        "output": "json"
    }
    if search_types:
        url_params["types"] = search_types
    
    base_url = "https://restapi.amap.com/v3/place/around"
    query_string_parts = []
    for k, v in url_params.items():
        if v: # Only add if value is not empty or None
            query_string_parts.append(f"{k}={requests.utils.quote(str(v))}")

    url = f"{base_url}?{'&'.join(query_string_parts)}"
    
    pois = []
    response_text = "N/A"
    try:
        response = requests.get(url, timeout=10)
        response_text = response.text
        response.raise_for_status()
        data = response.json()
        
        if data.get("status") == "1":
            if "pois" in data and data["pois"]:
                for poi_data in data["pois"]:
                    name = poi_data.get("name")
                    location_gcj_str = poi_data.get("location")
                    if name and location_gcj_str:
                        gcj_lon_str, gcj_lat_str = location_gcj_str.split(',')
                        wgs_lon, wgs_lat = gcj02_to_wgs84(float(gcj_lon_str), float(gcj_lat_str))
                        pois.append({
                            "name": name,
                            "wgs84_lon": wgs_lon,
                            "wgs84_lat": wgs_lat,
                            "gcj02_location_from_amap": location_gcj_str
                        })
                return None, pois
            else:
                return "未找到符合条件的POI", [] 
        else:
            return f"高德API错误: {data.get('info', '未知错误')} (infocode: {data.get('infocode')})", []
            
    except requests.exceptions.RequestException as e:
        return f"网络请求错误: {e}", []
    except json.JSONDecodeError:
        return f"解析高德API响应失败. URL: {url}, Response: {response_text}", []
    except Exception as e:
        return f"处理POI数据时发生未知错误: {e}", []

# --- KML Helper Functions ---
def create_kml_placemark(name, lon, lat, description=""):
    placemark = ET.Element("Placemark")
    ET.SubElement(placemark, "name").text = name
    if description:
        ET.SubElement(placemark, "description").text = description
    point = ET.SubElement(placemark, "Point")
    ET.SubElement(point, "coordinates").text = f"{lon},{lat},0"
    return placemark

def create_kml_circle_placemark(name, center_lon, center_lat, radius_meters, description=""):
    """创建圆形KML Placemark，使用Polygon近似圆形"""
    placemark = ET.Element("Placemark")
    ET.SubElement(placemark, "name").text = name
    if description:
        ET.SubElement(placemark, "description").text = description
    
    # 创建圆形的多边形近似（36个点）
    polygon = ET.SubElement(placemark, "Polygon")
    outer_boundary = ET.SubElement(polygon, "outerBoundaryIs")
    linear_ring = ET.SubElement(outer_boundary, "LinearRing")
    
    # 计算圆周上的点
    coordinates_list = []
    num_points = 36  # 圆周上的点数
    
    # 地球半径（米）
    earth_radius = 6378137.0
    
    for i in range(num_points + 1):  # +1 to close the polygon
        angle = 2 * math.pi * i / num_points
        
        # 计算相对于中心点的偏移（以度为单位）
        lat_offset = (radius_meters * math.cos(angle)) / earth_radius * (180 / math.pi)
        lon_offset = (radius_meters * math.sin(angle)) / (earth_radius * math.cos(math.radians(center_lat))) * (180 / math.pi)
        
        point_lat = center_lat + lat_offset
        point_lon = center_lon + lon_offset
        
        coordinates_list.append(f"{point_lon},{point_lat},0")
    
    coordinates_text = " ".join(coordinates_list)
    ET.SubElement(linear_ring, "coordinates").text = coordinates_text
    
    return placemark

def parse_kml_points(file_path):
    """解析KML文件中的点信息"""
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        # 确定命名空间
        namespace = ''
        if '}' in root.tag:
            namespace = root.tag.split('}')[0] + '}'
        
        points = []
        placemark_query = f'.//{namespace}Placemark'
        
        for placemark in root.findall(placemark_query):
            # 获取名称
            name_elem = placemark.find(f'{namespace}name')
            if name_elem is None:
                name_elem = placemark.find(f'.//{namespace}name')
            name = name_elem.text.strip() if name_elem is not None and name_elem.text else "未命名点"
            
            # 获取坐标（只处理Point类型）
            point_elem = placemark.find(f'.//{namespace}Point')
            if point_elem is not None:
                coords_elem = point_elem.find(f'{namespace}coordinates')
                if coords_elem is None:
                    coords_elem = point_elem.find(f'.//{namespace}coordinates')
                
                if coords_elem is not None and coords_elem.text:
                    coords_text = coords_elem.text.strip()
                    try:
                        # 处理坐标格式：lon,lat,alt 或 lon,lat
                        coords_parts = coords_text.replace(',', ' ').split()
                        lon = float(coords_parts[0])
                        lat = float(coords_parts[1])
                        
                        # 获取描述
                        desc_elem = placemark.find(f'{namespace}description')
                        if desc_elem is None:
                            desc_elem = placemark.find(f'.//{namespace}description')
                        description = desc_elem.text.strip() if desc_elem is not None and desc_elem.text else ""
                        
                        points.append({
                            'name': name,
                            'lon': lon,
                            'lat': lat,
                            'description': description
                        })
                    except (ValueError, IndexError):
                        continue
        
        return points, None
    except ET.ParseError as e:
        return [], f"KML文件解析错误: {e}"
    except Exception as e:
        return [], f"读取KML文件时发生错误: {e}"

def pretty_print_xml(xml_string):
    parsed_string = parseString(xml_string)
    return parsed_string.toprettyxml(indent="  ")

# --- GUI 主逻辑 ---
class GeoSpatialApp:
    def __init__(self, master):
        self.master = master
        # master.title("地理空间工具集") # Updated title - Removed, as Frame doesn't have title
        # master.geometry("800x700") # Adjusted initial size - Removed, as Frame doesn't have geometry

        # API Key Check
        if not config.get_amap_api_key():
            messagebox.showwarning("API Key未配置", 
                                 "请在配置中设置您的高德Web服务API Key。\n"
                                 "您可以从这里申请：https://lbs.amap.com/dev/key/app\n"
                                 "点击'配置管理'按钮进行设置。")

        # Main PanedWindow for layout
        main_paned_window = ttk.PanedWindow(master, orient=tk.VERTICAL)
        main_paned_window.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Top Frame for POI Search
        poi_search_frame_container = ttk.Frame(main_paned_window)
        main_paned_window.add(poi_search_frame_container, weight=1)

        # Input Frame for POI Search
        input_frame = ttk.LabelFrame(poi_search_frame_container, text="周边POI查询 (WGS-84输入)")
        input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        # 配置管理按钮
        config_frame = ttk.Frame(input_frame)
        config_frame.grid(row=0, column=0, columnspan=3, pady=5, sticky="ew")
        
        self.config_button = ttk.Button(config_frame, text="配置管理", command=self.open_config_dialog)
        self.config_button.pack(side=tk.LEFT, padx=5)
        
        # 收藏位置下拉框
        ttk.Label(config_frame, text="收藏位置:").pack(side=tk.LEFT, padx=(20, 5))
        self.favorite_var = tk.StringVar()
        self.favorite_combo = ttk.Combobox(config_frame, textvariable=self.favorite_var, width=15, state="readonly")
        self.favorite_combo.pack(side=tk.LEFT, padx=5)
        self.favorite_combo.bind("<<ComboboxSelected>>", self.on_favorite_selected)
        self.update_favorite_locations()
        
        ttk.Label(input_frame, text="WGS-84 经度:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.lon_entry = ttk.Entry(input_frame, width=25)
        self.lon_entry.grid(row=1, column=1, padx=5, pady=5)
        self.lon_entry.insert(0, "119.429737") 
        
        # 添加收藏按钮
        self.add_favorite_button = ttk.Button(input_frame, text="收藏", command=self.add_current_location_to_favorites)
        self.add_favorite_button.grid(row=1, column=2, padx=5, pady=5)

        ttk.Label(input_frame, text="WGS-84 纬度:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.lat_entry = ttk.Entry(input_frame, width=25)
        self.lat_entry.grid(row=2, column=1, padx=5, pady=5)
        self.lat_entry.insert(0, "25.97546") 

        ttk.Label(input_frame, text="查询关键字:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.keyword_entry = ttk.Entry(input_frame, width=25)
        self.keyword_entry.grid(row=3, column=1, padx=5, pady=5)
        self.keyword_entry.insert(0, "加油站") 

        ttk.Label(input_frame, text="查询半径 (km):").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.radius_entry = ttk.Entry(input_frame, width=25)
        self.radius_entry.grid(row=4, column=1, padx=5, pady=5)
        self.radius_entry.insert(0, "6") 

        self.search_button = ttk.Button(input_frame, text="查询POI", command=self.perform_search)
        self.search_button.grid(row=5, column=0, columnspan=3, pady=10)
        
        # 配置列权重
        input_frame.grid_columnconfigure(1, weight=1)

        # Results Frame for POI Search
        results_frame = ttk.LabelFrame(poi_search_frame_container, text="查询结果")
        results_frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        poi_search_frame_container.grid_rowconfigure(1, weight=1)
        poi_search_frame_container.grid_columnconfigure(0, weight=1)

        columns = ("name", "lon", "lat")
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show="headings")
        self.results_tree.heading("name", text="名称")
        self.results_tree.heading("lon", text="WGS-84 经度")
        self.results_tree.heading("lat", text="WGS-84 纬度")

        self.results_tree.column("name", width=250, stretch=tk.YES)
        self.results_tree.column("lon", width=150, anchor="e", stretch=tk.YES)
        self.results_tree.column("lat", width=150, anchor="e", stretch=tk.YES)
        
        vsb = ttk.Scrollbar(results_frame, orient="vertical", command=self.results_tree.yview)
        hsb = ttk.Scrollbar(results_frame, orient="horizontal", command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.results_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)
        
        # Export buttons for POI Search results
        export_buttons_frame = ttk.Frame(poi_search_frame_container)
        export_buttons_frame.grid(row=2, column=0, pady=(5,10), sticky="ew", padx=10)

        self.export_excel_button = ttk.Button(export_buttons_frame, text="导出Excel", command=self.export_to_excel)
        self.export_excel_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        self.export_kml_button = ttk.Button(export_buttons_frame, text="导出KML", command=self.export_to_kml)
        self.export_kml_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Bottom Frame for Conversion Tools
        conversion_tools_frame_container = ttk.Frame(main_paned_window)
        main_paned_window.add(conversion_tools_frame_container, weight=0) # Less weight initially

        tools_frame = ttk.LabelFrame(conversion_tools_frame_container, text="格式与坐标转换工具 (Excel/KML, WGS-84)")
        tools_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.excel_to_kml_button = ttk.Button(tools_frame, text="Excel 转 KML", command=self.convert_excel_to_kml)
        self.excel_to_kml_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        self.kml_to_excel_button = ttk.Button(tools_frame, text="KML 转 Excel", command=self.convert_kml_to_excel)
        self.kml_to_excel_button.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self.address_to_coords_button = ttk.Button(tools_frame, text="地址 转 经纬度 (Excel导入/导出)", command=self.convert_address_to_coords_excel)
        self.address_to_coords_button.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        self.coords_to_address_button = ttk.Button(tools_frame, text="经纬度 转 地址 (Excel导入/导出)", command=self.convert_coords_to_address_excel)
        self.coords_to_address_button.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        # 点画圆功能按钮
        self.point_to_circle_button = ttk.Button(tools_frame, text="KML点画圆 (上传KML点文件生成圆形)", command=self.convert_points_to_circles)
        self.point_to_circle_button.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
        
        tools_frame.grid_columnconfigure(0, weight=1)
        tools_frame.grid_columnconfigure(1, weight=1)

        # Status bar for this specific tool, placed within its own frame (master)
        self.tool_status_label = ttk.Label(self.master, text="地理空间工具准备就绪。") 
        self.tool_status_label.pack(side=tk.BOTTOM, pady=5, padx=10, fill=tk.X)
        
        self.current_results = []

    def perform_search(self):
        self.tool_status_label.config(text="正在查询...") # Use the new status label
        self.master.update_idletasks()

        if not config.get_amap_api_key():
            messagebox.showerror("API Key错误", "请先在配置中设置有效的高德API Key。")
            self.tool_status_label.config(text="API Key未配置，请在配置管理中设置。") # Use the new status label
            return

        try:
            wgs_lon_str = self.lon_entry.get().strip()
            wgs_lat_str = self.lat_entry.get().strip()
            keywords = self.keyword_entry.get().strip()
            radius_km_str = self.radius_entry.get().strip()

            if not all([wgs_lon_str, wgs_lat_str, keywords, radius_km_str]):
                messagebox.showerror("输入错误", "所有输入字段均不能为空。")
                self.tool_status_label.config(text="输入字段不能为空。") # Use the new status label
                return

            wgs_lon = float(wgs_lon_str)
            wgs_lat = float(wgs_lat_str)
            radius_km = float(radius_km_str)
        except ValueError:
            messagebox.showerror("输入错误", "经纬度和半径必须是有效的数字。")
            self.tool_status_label.config(text="输入错误：经纬度和半径需为数字。") # Use the new status label
            return

        if not keywords:
            messagebox.showerror("输入错误", "查询关键字不能为空。") # Redundant due to check above, but good to keep
            self.tool_status_label.config(text="关键字不能为空。") # Use the new status label
            return
        
        if radius_km <= 0:
            messagebox.showerror("输入错误", "查询半径必须大于0。")
            self.tool_status_label.config(text="半径错误：必须大于0。") # Use the new status label
            return

        radius_meters = radius_km * 1000

        for i in self.results_tree.get_children():
            self.results_tree.delete(i)
        self.current_results = []

        # WGS-84 to GCJ-02 for API call is handled within search_nearby_pois_amap
        error_msg, pois = search_nearby_pois_amap(wgs_lon, wgs_lat, radius_meters, keywords)

        if error_msg:
            messagebox.showerror("查询失败", f"{error_msg}")
            self.tool_status_label.config(text=f"查询失败: {error_msg}") # Use the new status label
        elif pois:
            self.current_results = pois
            for poi in pois:
                self.results_tree.insert("", "end", values=(
                    poi["name"],
                    f"{poi['wgs84_lon']:.6f}", 
                    f"{poi['wgs84_lat']:.6f}"
                ))
            self.tool_status_label.config(text=f"查询完成，找到 {len(pois)} 个结果。") # Use the new status label
        else:
            messagebox.showinfo("查询结果", "未找到符合条件的POI。")
            self.tool_status_label.config(text="未找到结果。") # Use the new status label

    def export_to_excel(self):
        if not self.current_results:
            messagebox.showinfo("无数据", "没有可导出的查询结果。")
            return

        try:
            # Try to import openpyxl here to catch if it's not installed
            import openpyxl
        except ImportError:
            messagebox.showerror("缺少库", "需要安装 openpyxl 库才能导出Excel。\n请运行: pip install openpyxl")
            self.tool_status_label.config(text="导出失败:缺少openpyxl库。")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
            title="保存查询结果为Excel"
        )

        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "查询结果"

            sheet["A1"] = "名称"
            sheet["B1"] = "WGS-84 经度"
            sheet["C1"] = "WGS-84 纬度"

            for row_idx, poi in enumerate(self.current_results, start=2):
                sheet[f"A{row_idx}"] = poi["name"]
                sheet[f"B{row_idx}"] = poi["wgs84_lon"]
                sheet[f"C{row_idx}"] = poi["wgs84_lat"]
            
            workbook.save(file_path)
            messagebox.showinfo("导出成功", f"结果已成功导出到:\n{file_path}")
            self.tool_status_label.config(text=f"结果已导出到 {file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出到Excel时发生错误: {e}")
            self.tool_status_label.config(text=f"导出失败: {e}")

    def export_to_kml(self):
        if not self.current_results:
            messagebox.showinfo("无数据", "没有可导出的查询结果。")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".kml",
            filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")],
            title="保存查询结果为KML"
        )

        if not file_path:
            return

        kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
        document = ET.SubElement(kml, "Document")

        for poi in self.current_results:
            placemark = create_kml_placemark(poi["name"], poi["wgs84_lon"], poi["wgs84_lat"])
            document.append(placemark)

        try:
            tree_string = ET.tostring(kml, encoding='utf-8', method='xml')
            pretty_xml_string = pretty_print_xml(tree_string.decode('utf-8'))
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(pretty_xml_string)
            messagebox.showinfo("导出成功", f"结果已成功导出到:\n{file_path}")
            self.tool_status_label.config(text=f"KML结果已导出到 {file_path}")
        except Exception as e:
            messagebox.showerror("导出KML失败", f"导出到KML时发生错误: {e}")
            self.tool_status_label.config(text=f"导出KML失败: {e}")

    def convert_excel_to_kml(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件进行转换",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
            document = ET.SubElement(kml, "Document")

            # Assuming header in the first row
            # A: 名称, B: 经度, C: 纬度, D: 备注
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 3 or not all(row[:3]): # Need at least name, lon, lat
                    self.tool_status_label.config(text=f"跳过Excel第 {row_idx} 行：名称、经纬度不能为空。")
                    continue
                name = str(row[0])
                try:
                    lon = float(row[1])
                    lat = float(row[2])
                except (ValueError, TypeError):
                    self.tool_status_label.config(text=f"跳过Excel第 {row_idx} 行：经纬度格式错误。")
                    continue
                
                description = str(row[3]) if len(row) > 3 and row[3] is not None else ""
                placemark = create_kml_placemark(name, lon, lat, description)
                document.append(placemark)
            
            if not list(document):
                messagebox.showinfo("无数据转换", "Excel文件中没有有效数据可转换为KML。")
                self.tool_status_label.config(text="Excel转KML：无有效数据。")
                return

            save_path = filedialog.asksaveasfilename(
                defaultextension=".kml",
                filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")],
                title="保存KML文件"
            )
            if not save_path:
                return

            tree_string = ET.tostring(kml, encoding='utf-8', method='xml')
            pretty_xml_string = pretty_print_xml(tree_string.decode('utf-8'))
            with open(save_path, "w", encoding="utf-8") as f:
                f.write(pretty_xml_string)
            messagebox.showinfo("转换成功", f"Excel已成功转换为KML:\n{save_path}")
            self.tool_status_label.config(text=f"Excel已转为KML: {save_path}")

        except FileNotFoundError:
            messagebox.showerror("文件错误", f"未找到文件: {file_path}")
            self.tool_status_label.config(text=f"Excel转KML错误: 文件未找到")
        except openpyxl.utils.exceptions.InvalidFileException:
            messagebox.showerror("文件错误", f"无法打开或不是有效的Excel文件: {file_path}")
            self.tool_status_label.config(text=f"Excel转KML错误: 无效Excel文件")
        except Exception as e:
            messagebox.showerror("转换失败", f"Excel转KML时发生错误: {e}")
            self.tool_status_label.config(text=f"Excel转KML错误: {e}")

    def convert_kml_to_excel(self):
        file_path = filedialog.askopenfilename(
            title="选择KML文件进行转换",
            filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Determine the namespace from the root element
            namespace = ''
            if '}' in root.tag:
                namespace = root.tag.split('}')[0] + '}' # e.g. {http://www.opengis.net/kml/2.2}
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "KML数据"
            sheet["A1"] = "名称"
            sheet["B1"] = "WGS-84 经度"
            sheet["C1"] = "WGS-84 纬度"
            sheet["D1"] = "备注信息"

            row_num = 2
            # Find all Placemark elements, regardless of their depth in the document
            # Construct XPath queries with the determined namespace
            placemark_query = f'.//{namespace}Placemark'
            name_query = f'.//{namespace}name'
            coordinates_query = f'.//{namespace}Point/{namespace}coordinates' # More specific path to coordinates
            description_query = f'.//{namespace}description'

            for placemark in root.findall(placemark_query):
                name_elem = placemark.find(name_query.split('/')[-1]) # Find direct child 'name'
                if name_elem is None: # Try finding it anywhere under placemark if not direct child
                    name_elem = placemark.find(f'.//{namespace}name') 
                name = name_elem.text.strip() if name_elem is not None and name_elem.text else "未命名"
                
                # Try specific path first, then more general
                coords_elem = placemark.find(coordinates_query.split('/')[-2] + '/' + coordinates_query.split('/')[-1]) 
                if coords_elem is None: # If not under Point, try finding coordinates anywhere under placemark
                     coords_elem = placemark.find(f'.//{namespace}coordinates')

                if coords_elem is None or not coords_elem.text:
                    self.tool_status_label.config(text=f"跳过Placemark '{name}': 缺少坐标。")
                    continue
                
                coords_text = coords_elem.text.strip()
                try:
                    # Handle coordinates that might have altitude or be space-separated
                    coords_parts = coords_text.replace(',', ' ').split()
                    lon = float(coords_parts[0])
                    lat = float(coords_parts[1])
                except (ValueError, IndexError):
                    self.tool_status_label.config(text=f"跳过Placemark '{name}': 坐标格式错误 '{coords_text}'。")
                    continue

                desc_elem = placemark.find(description_query.split('/')[-1]) # Find direct child 'description'
                if desc_elem is None: # Try finding it anywhere under placemark if not direct child
                    desc_elem = placemark.find(f'.//{namespace}description')
                description = desc_elem.text.strip() if desc_elem is not None and desc_elem.text else ""

                sheet[f"A{row_num}"] = name
                sheet[f"B{row_num}"] = lon
                sheet[f"C{row_num}"] = lat
                sheet[f"D{row_num}"] = description
                row_num += 1
            
            if row_num == 2:
                messagebox.showinfo("无数据转换", "KML文件中没有找到可转换的Placemark数据。")
                self.tool_status_label.config(text="KML转Excel：无有效数据。")
                return

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
                title="保存Excel文件"
            )
            if not save_path:
                return

            workbook.save(save_path)
            messagebox.showinfo("转换成功", f"KML已成功转换为Excel:\n{save_path}")
            self.tool_status_label.config(text=f"KML已转为Excel: {save_path}")

        except FileNotFoundError:
            messagebox.showerror("文件错误", f"未找到文件: {file_path}")
            self.tool_status_label.config(text=f"KML转Excel错误: 文件未找到")
        except ET.ParseError:
            messagebox.showerror("文件错误", f"无法解析KML文件 (格式可能无效): {file_path}")
            self.tool_status_label.config(text=f"KML转Excel错误: 无效KML文件")
        except Exception as e:
            messagebox.showerror("转换失败", f"KML转Excel时发生错误: {e}")
            self.tool_status_label.config(text=f"KML转Excel错误: {e}")

    def _process_excel_coords_conversion(self, conversion_type):
        """Helper for Address to Coords and Coords to Address Excel processing."""
        input_file_path = filedialog.askopenfilename(
            title=f"选择包含 {'地址' if conversion_type == 'addr_to_coord' else '经纬度'} 的Excel文件",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not input_file_path:
            return

        if not config.get_amap_api_key():
            messagebox.showerror("API Key错误", "请先在配置中设置有效的高德API Key。")
            self.tool_status_label.config(text="API Key未配置，请在配置管理中设置。") # Use the new status label
            return

        try:
            workbook = openpyxl.load_workbook(input_file_path)
            sheet = workbook.active
            output_data = []
            headers = [cell.value for cell in sheet[1]] # Get headers from first row
            output_data.append(headers)

            # Determine column indices (more robust than fixed A, B, C)
            try:
                addr_col_idx = headers.index("地址") if "地址" in headers else 0
                lon_col_idx = headers.index("经度") if "经度" in headers else 1
                lat_col_idx = headers.index("纬度") if "纬度" in headers else 2
            except ValueError:
                 messagebox.showerror("表头错误", "Excel文件必须包含'地址', '经度', '纬度'的表头。")
                 self.tool_status_label.config(text="Excel表头错误。")
                 return

            self.tool_status_label.config(text="正在处理Excel文件...")
            self.master.update_idletasks()
            processed_count = 0

            for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
                row_values = [cell.value for cell in row_cells]
                # Ensure row has enough columns based on header length
                if len(row_values) < len(headers):
                    row_values.extend([None] * (len(headers) - len(row_values)))
                
                current_row_output = list(row_values) # Start with original data

                if conversion_type == 'addr_to_coord':
                    address = str(row_values[addr_col_idx]).strip() if row_values[addr_col_idx] else None
                    city = "" # Attempt without city first, or prompt user, or add a city column
                    if address:
                        err, lon_wgs, lat_wgs = get_coords_from_amap(address, city)
                        if err:
                            self.tool_status_label.config(text=f"第{row_idx}行 '{address}': {err}")
                        elif lon_wgs is not None and lat_wgs is not None:
                            current_row_output[lon_col_idx] = lon_wgs
                            current_row_output[lat_col_idx] = lat_wgs
                            processed_count += 1
                        else:
                             self.tool_status_label.config(text=f"第{row_idx}行 '{address}': 未找到坐标。")
                    else:
                        self.tool_status_label.config(text=f"第{row_idx}行: 地址为空，跳过。")
                
                elif conversion_type == 'coord_to_addr':
                    try:
                        lon_wgs = float(row_values[lon_col_idx]) if row_values[lon_col_idx] is not None else None
                        lat_wgs = float(row_values[lat_col_idx]) if row_values[lat_col_idx] is not None else None
                    except (ValueError, TypeError):
                        self.tool_status_label.config(text=f"第{row_idx}行: 经纬度格式错误，跳过。")
                        lon_wgs, lat_wgs = None, None

                    if lon_wgs is not None and lat_wgs is not None:
                        address_result = get_address_from_amap(lon_wgs, lat_wgs)
                        current_row_output[addr_col_idx] = address_result
                        if not address_result.startswith("错误") and not address_result.startswith("高德API错误") and not address_result.startswith("网络请求错误") and not address_result.startswith("解析高德API响应失败") and address_result != "地址未找到":
                            processed_count +=1
                        else:
                            self.tool_status_label.config(text=f"第{row_idx}行 ({lon_wgs},{lat_wgs}): {address_result}")
                    else:
                        self.tool_status_label.config(text=f"第{row_idx}行: 经纬度为空，跳过。")
                
                output_data.append(current_row_output)
                if row_idx % 10 == 0: # Update status periodically
                    self.master.update_idletasks()
            
            if processed_count == 0 and len(list(sheet.iter_rows(min_row=2))) > 0:
                 messagebox.showwarning("处理完成", "未成功处理任何行。请检查输入数据和API Key。")
                 self.tool_status_label.config(text="未成功处理任何行。")
                 return
            elif processed_count == 0:
                 messagebox.showinfo("无数据", "Excel文件中没有可处理的数据。")
                 self.tool_status_label.config(text="Excel文件无数据。")
                 return

            output_file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
                title="保存处理结果到Excel"
            )
            if not output_file_path:
                return

            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            for r_idx, data_row in enumerate(output_data):
                for c_idx, value in enumerate(data_row):
                    new_sheet.cell(row=r_idx + 1, column=c_idx + 1, value=value)
            
            new_workbook.save(output_file_path)
            messagebox.showinfo("处理成功", f"文件已处理并保存到:\n{output_file_path}\n成功处理 {processed_count} 行。")
            self.tool_status_label.config(text=f"Excel处理完成: {output_file_path}")

        except FileNotFoundError:
            messagebox.showerror("文件错误", f"未找到文件: {input_file_path}")
            self.tool_status_label.config(text=f"Excel处理错误: 文件未找到")
        except openpyxl.utils.exceptions.InvalidFileException:
            messagebox.showerror("文件错误", f"无法打开或不是有效的Excel文件: {input_file_path}")
            self.tool_status_label.config(text=f"Excel处理错误: 无效Excel文件")
        except Exception as e:
            messagebox.showerror("处理失败", f"处理Excel文件时发生错误: {e}")
            self.tool_status_label.config(text=f"Excel处理错误: {e}")

    def convert_address_to_coords_excel(self):
        self._process_excel_coords_conversion('addr_to_coord')

    def convert_coords_to_address_excel(self):
        self._process_excel_coords_conversion('coord_to_addr')

    def convert_points_to_circles(self):
        """KML点画圆功能：上传KML点文件，输入半径，生成圆形KML文件"""
        # 选择输入的KML点文件
        input_file_path = filedialog.askopenfilename(
            title="选择包含点信息的KML文件",
            filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")]
        )
        if not input_file_path:
            return

        # 解析KML文件中的点
        self.tool_status_label.config(text="正在解析KML文件...")
        self.master.update_idletasks()
        
        points, error_msg = parse_kml_points(input_file_path)
        if error_msg:
            messagebox.showerror("文件解析失败", error_msg)
            self.tool_status_label.config(text=f"KML解析失败: {error_msg}")
            return
        
        if not points:
            messagebox.showinfo("无点数据", "KML文件中没有找到有效的点数据。")
            self.tool_status_label.config(text="KML文件中无有效点数据。")
            return
        
        # 创建对话框让用户输入圆半径
        radius_dialog = tk.Toplevel(self.master)
        radius_dialog.title("设置圆半径")
        radius_dialog.geometry("400x200")
        radius_dialog.transient(self.master)
        radius_dialog.grab_set()
        
        # 居中显示对话框
        radius_dialog.update_idletasks()
        x = (radius_dialog.winfo_screenwidth() // 2) - (radius_dialog.winfo_width() // 2)
        y = (radius_dialog.winfo_screenheight() // 2) - (radius_dialog.winfo_height() // 2)
        radius_dialog.geometry(f"+{x}+{y}")
        
        # 对话框内容
        info_frame = ttk.Frame(radius_dialog)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(info_frame, text=f"已找到 {len(points)} 个点", font=("Arial", 12)).pack(pady=(0, 10))
        ttk.Label(info_frame, text="请输入圆的半径（米）:", font=("Arial", 10)).pack(pady=(0, 5))
        
        radius_var = tk.StringVar(value="1000")  # 默认1000米
        radius_entry = ttk.Entry(info_frame, textvariable=radius_var, font=("Arial", 10), width=20)
        radius_entry.pack(pady=(0, 15))
        radius_entry.focus()
        
        # 按钮框架
        button_frame = ttk.Frame(info_frame)
        button_frame.pack(fill=tk.X)
        
        result_radius = [None]  # 使用列表来存储结果，以便在内部函数中修改
        
        def on_ok():
            try:
                radius = float(radius_var.get().strip())
                if radius <= 0:
                    messagebox.showerror("输入错误", "半径必须大于0")
                    return
                result_radius[0] = radius
                radius_dialog.destroy()
            except ValueError:
                messagebox.showerror("输入错误", "请输入有效的数字")
        
        def on_cancel():
            radius_dialog.destroy()
        
        ttk.Button(button_frame, text="确定", command=on_ok).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="取消", command=on_cancel).pack(side=tk.LEFT)
        
        # 绑定回车键
        radius_entry.bind('<Return>', lambda e: on_ok())
        
        # 等待对话框关闭
        self.master.wait_window(radius_dialog)
        
        # 检查用户是否取消了操作
        if result_radius[0] is None:
            self.tool_status_label.config(text="操作已取消。")
            return
        
        radius_meters = result_radius[0]
        
        # 选择输出文件路径
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".kml",
            filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")],
            title="保存圆形KML文件"
        )
        if not output_file_path:
            self.tool_status_label.config(text="未选择输出文件，操作取消。")
            return
        
        # 生成圆形KML文件
        try:
            self.tool_status_label.config(text="正在生成圆形KML文件...")
            self.master.update_idletasks()
            
            kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
            document = ET.SubElement(kml, "Document")
            
            # 添加文档信息
            ET.SubElement(document, "name").text = "点画圆结果"
            ET.SubElement(document, "description").text = f"基于 {len(points)} 个点生成的圆形，半径: {radius_meters} 米"
            
            # 为每个点创建圆形
            for i, point in enumerate(points):
                circle_name = f"{point['name']}_圆形_{radius_meters}m"
                circle_description = f"原点: {point['name']}\n半径: {radius_meters} 米"
                if point['description']:
                    circle_description += f"\n原描述: {point['description']}"
                
                circle_placemark = create_kml_circle_placemark(
                    circle_name, 
                    point['lon'], 
                    point['lat'], 
                    radius_meters, 
                    circle_description
                )
                document.append(circle_placemark)
                
                # 更新进度
                if (i + 1) % 10 == 0 or i == len(points) - 1:
                    self.tool_status_label.config(text=f"正在生成圆形... ({i + 1}/{len(points)})")
                    self.master.update_idletasks()
            
            # 保存KML文件
            tree_string = ET.tostring(kml, encoding='utf-8', method='xml')
            pretty_xml_string = pretty_print_xml(tree_string.decode('utf-8'))
            
            with open(output_file_path, "w", encoding="utf-8") as f:
                f.write(pretty_xml_string)
            
            messagebox.showinfo(
                "生成成功", 
                f"圆形KML文件已成功生成！\n\n"
                f"输入点数: {len(points)}\n"
                f"圆半径: {radius_meters} 米\n"
                f"输出文件: {output_file_path}"
            )
            self.tool_status_label.config(text=f"圆形KML已生成: {output_file_path}")
            
        except Exception as e:
            messagebox.showerror("生成失败", f"生成圆形KML文件时发生错误: {e}")
            self.tool_status_label.config(text=f"生成圆形KML失败: {e}")
    
    def open_config_dialog(self):
        """打开配置管理对话框"""
        config_window = tk.Toplevel(self.master)
        config_window.title("配置管理")
        config_window.geometry("500x400")
        config_window.transient(self.master)
        config_window.grab_set()
        
        # API Key配置
        api_frame = ttk.LabelFrame(config_window, text="高德地图API配置")
        api_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(api_frame, text="API Key:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        api_key_var = tk.StringVar(value=config.get_amap_api_key())
        api_key_entry = ttk.Entry(api_frame, textvariable=api_key_var, width=50, show="*")
        api_key_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        def toggle_api_key_visibility():
            if api_key_entry.cget('show') == '*':
                api_key_entry.config(show='')
                show_button.config(text="隐藏")
            else:
                api_key_entry.config(show='*')
                show_button.config(text="显示")
        
        show_button = ttk.Button(api_frame, text="显示", command=toggle_api_key_visibility)
        show_button.grid(row=0, column=2, padx=5, pady=5)
        
        api_frame.grid_columnconfigure(1, weight=1)
        
        # 收藏位置管理
        favorites_frame = ttk.LabelFrame(config_window, text="收藏位置管理")
        favorites_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 收藏位置列表
        columns = ("name", "lng", "lat")
        favorites_tree = ttk.Treeview(favorites_frame, columns=columns, show="headings", height=8)
        favorites_tree.heading("name", text="名称")
        favorites_tree.heading("lng", text="经度")
        favorites_tree.heading("lat", text="纬度")
        
        favorites_tree.column("name", width=150)
        favorites_tree.column("lng", width=120)
        favorites_tree.column("lat", width=120)
        
        favorites_scrollbar = ttk.Scrollbar(favorites_frame, orient="vertical", command=favorites_tree.yview)
        favorites_tree.configure(yscrollcommand=favorites_scrollbar.set)
        
        favorites_tree.grid(row=0, column=0, columnspan=4, sticky="nsew", padx=5, pady=5)
        favorites_scrollbar.grid(row=0, column=4, sticky="ns", pady=5)
        
        def refresh_favorites_tree():
            for item in favorites_tree.get_children():
                favorites_tree.delete(item)
            for loc in config.get_favorite_locations():
                favorites_tree.insert("", "end", values=(loc['name'], f"{loc['lng']:.6f}", f"{loc['lat']:.6f}"))
        
        refresh_favorites_tree()
        
        # 添加新收藏位置
        add_frame = ttk.Frame(favorites_frame)
        add_frame.grid(row=1, column=0, columnspan=5, sticky="ew", padx=5, pady=5)
        
        ttk.Label(add_frame, text="名称:").grid(row=0, column=0, padx=5)
        name_var = tk.StringVar()
        name_entry = ttk.Entry(add_frame, textvariable=name_var, width=15)
        name_entry.grid(row=0, column=1, padx=5)
        
        ttk.Label(add_frame, text="经度:").grid(row=0, column=2, padx=5)
        lng_var = tk.StringVar()
        lng_entry = ttk.Entry(add_frame, textvariable=lng_var, width=12)
        lng_entry.grid(row=0, column=3, padx=5)
        
        ttk.Label(add_frame, text="纬度:").grid(row=0, column=4, padx=5)
        lat_var = tk.StringVar()
        lat_entry = ttk.Entry(add_frame, textvariable=lat_var, width=12)
        lat_entry.grid(row=0, column=5, padx=5)
        
        def add_favorite():
            name = name_var.get().strip()
            try:
                lng = float(lng_var.get().strip())
                lat = float(lat_var.get().strip())
                if name and config.add_favorite_location(name, lng, lat):
                    refresh_favorites_tree()
                    self.update_favorite_locations()
                    name_var.set("")
                    lng_var.set("")
                    lat_var.set("")
                    messagebox.showinfo("成功", f"已添加收藏位置: {name}")
                else:
                    messagebox.showerror("错误", "请输入有效的名称和坐标")
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数字坐标")
        
        def delete_favorite():
            selected = favorites_tree.selection()
            if selected:
                item = favorites_tree.item(selected[0])
                name = item['values'][0]
                if messagebox.askyesno("确认删除", f"确定要删除收藏位置 '{name}' 吗？"):
                    if config.remove_favorite_location(name):
                        refresh_favorites_tree()
                        self.update_favorite_locations()
                        messagebox.showinfo("成功", f"已删除收藏位置: {name}")
            else:
                messagebox.showwarning("提示", "请先选择要删除的收藏位置")
        
        add_button = ttk.Button(add_frame, text="添加", command=add_favorite)
        add_button.grid(row=0, column=6, padx=10)
        
        delete_button = ttk.Button(add_frame, text="删除选中", command=delete_favorite)
        delete_button.grid(row=0, column=7, padx=5)
        
        favorites_frame.grid_rowconfigure(0, weight=1)
        favorites_frame.grid_columnconfigure(0, weight=1)
        
        # 保存按钮
        button_frame = ttk.Frame(config_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def save_config():
            # 保存API Key
            if config.set_amap_api_key(api_key_var.get().strip()):
                messagebox.showinfo("成功", "配置已保存")
                config_window.destroy()
            else:
                messagebox.showerror("错误", "保存配置失败")
        
        def cancel_config():
            config_window.destroy()
        
        ttk.Button(button_frame, text="保存", command=save_config).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="取消", command=cancel_config).pack(side=tk.RIGHT, padx=5)
    
    def update_favorite_locations(self):
        """更新收藏位置下拉框"""
        favorites = config.get_favorite_locations()
        favorite_names = [loc['name'] for loc in favorites]
        self.favorite_combo['values'] = favorite_names
        if favorite_names:
            self.favorite_combo.set('')  # 清空选择
    
    def on_favorite_selected(self, event):
        """当选择收藏位置时，自动填充坐标"""
        selected_name = self.favorite_var.get()
        if selected_name:
            favorites = config.get_favorite_locations()
            for loc in favorites:
                if loc['name'] == selected_name:
                    self.lon_entry.delete(0, tk.END)
                    self.lon_entry.insert(0, str(loc['lng']))
                    self.lat_entry.delete(0, tk.END)
                    self.lat_entry.insert(0, str(loc['lat']))
                    break
    
    def add_current_location_to_favorites(self):
        """将当前输入的坐标添加到收藏位置"""
        try:
            lng = float(self.lon_entry.get().strip())
            lat = float(self.lat_entry.get().strip())
            
            # 弹出对话框让用户输入名称
            name = tk.simpledialog.askstring("添加收藏位置", "请输入位置名称:")
            if name and name.strip():
                if config.add_favorite_location(name.strip(), lng, lat):
                    self.update_favorite_locations()
                    messagebox.showinfo("成功", f"已添加收藏位置: {name.strip()}")
                else:
                    messagebox.showerror("错误", "添加收藏位置失败")
        except ValueError:
            messagebox.showerror("错误", "请输入有效的坐标")

if __name__ == "__main__":
    root = tk.Tk()
    app = GeoSpatialApp(root)
    root.mainloop()
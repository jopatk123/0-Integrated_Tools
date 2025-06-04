# -*- coding: utf-8 -*-
"""格式转换选项卡模块"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import xml.etree.ElementTree as ET
import openpyxl
from .amap_api import get_address_from_amap, get_coords_from_amap
from .kml_utils import create_kml_placemark, create_kml_circle_placemark, parse_kml_points, pretty_print_xml

class ConversionTab:
    """格式转换选项卡"""
    
    def __init__(self, parent, notebook, theme, config):
        self.parent = parent
        self.theme = theme
        self.config = config
        self.update_status = None  # 状态更新回调函数
        
        # 创建选项卡
        self.frame = ttk.Frame(notebook)
        notebook.add(self.frame, text="🔄 格式转换")
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置用户界面"""
        # 主容器
        main_frame = ttk.Frame(self.frame)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 标题
        title_label = ttk.Label(main_frame, text="格式与坐标转换工具 (Excel/KML, WGS-84)", 
                               font=("微软雅黑", 12, "bold"))
        title_label.pack(pady=(0, 20))
        
        # 创建按钮网格
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.BOTH, expand=True)
        
        # 第一行按钮
        row1_frame = ttk.Frame(button_frame)
        row1_frame.pack(fill=tk.X, pady=5)
        
        self.excel_to_kml_button = ttk.Button(row1_frame, text="Excel 转 KML", 
                                             command=self.convert_excel_to_kml)
        self.excel_to_kml_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        self.kml_to_excel_button = ttk.Button(row1_frame, text="KML 转 Excel", 
                                             command=self.convert_kml_to_excel)
        self.kml_to_excel_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 第二行按钮
        row2_frame = ttk.Frame(button_frame)
        row2_frame.pack(fill=tk.X, pady=5)
        
        self.address_to_coords_button = ttk.Button(row2_frame, text="地址 转 经纬度 (Excel导入/导出)", 
                                                  command=self.convert_address_to_coords_excel)
        self.address_to_coords_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        self.coords_to_address_button = ttk.Button(row2_frame, text="经纬度 转 地址 (Excel导入/导出)", 
                                                  command=self.convert_coords_to_address_excel)
        self.coords_to_address_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 第三行按钮
        row3_frame = ttk.Frame(button_frame)
        row3_frame.pack(fill=tk.X, pady=5)
        
        self.point_to_circle_button = ttk.Button(row3_frame, text="KML点画圆 (上传KML点文件生成圆形)", 
                                                command=self.convert_points_to_circles)
        self.point_to_circle_button.pack(fill=tk.X, expand=True)
        
        # 第四行按钮 - Excel模板下载（醒目提示）
        row4_frame = ttk.Frame(button_frame)
        row4_frame.pack(fill=tk.X, pady=10)  # 增加间距
        
        # 创建醒目的提示框架
        template_notice_frame = ttk.LabelFrame(row4_frame, text="⚠️ 重要提示")
        template_notice_frame.pack(fill=tk.X, pady=5)
        
        # 提示文本
        notice_text = ttk.Label(template_notice_frame, 
                               text="📋 使用前请先下载对应的Excel模板文件，确保数据格式正确！",
                               font=("微软雅黑", 10, "bold"),
                               foreground="#d63384")  # 醒目的红色
        notice_text.pack(pady=5)
        
        # 下载按钮 - 使用更醒目的样式
        self.download_excel_template_button = ttk.Button(template_notice_frame, 
                                                        text="📥 立即下载Excel模板", 
                                                        command=self.download_excel_templates)
        self.download_excel_template_button.pack(fill=tk.X, expand=True, pady=5)
        
        # 配置按钮样式使其更醒目
        style = ttk.Style()
        style.configure("Accent.TButton", font=("微软雅黑", 11, "bold"))
        self.download_excel_template_button.configure(style="Accent.TButton")
        
        # 说明文本
        info_frame = ttk.LabelFrame(main_frame, text="使用说明")
        info_frame.pack(fill=tk.X, pady=(20, 0))
        
        info_text = (
            "• Excel 转 KML: 将包含经纬度的Excel文件转换为KML格式\n"
            "• KML 转 Excel: 将KML文件中的点位信息导出为Excel\n"
            "• 地址转经纬度: 批量将地址转换为坐标（需要高德API）\n"
            "• 经纬度转地址: 批量将坐标转换为地址（需要高德API）\n"
            "• KML点画圆: 为KML文件中的每个点生成指定半径的圆形区域\n"
            "• 下载Excel模板: 提供标准格式的Excel模板文件，便于数据导入"
        )
        
        info_label = ttk.Label(info_frame, text=info_text, justify=tk.LEFT)
        info_label.pack(padx=10, pady=10)
    
    def convert_excel_to_kml(self):
        """Excel转KML"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件进行转换",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
            document = ET.SubElement(kml, "Document")
            
            # 假设第一行是标题，从第二行开始读取数据
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3 and row[0] and row[1] and row[2]:
                    try:
                        name = str(row[0])
                        lon = float(row[1])
                        lat = float(row[2])
                        description = str(row[3]) if len(row) > 3 and row[3] else ""
                        
                        placemark = create_kml_placemark(name, lon, lat, description)
                        document.append(placemark)
                    except (ValueError, TypeError):
                        continue
            
            # 保存KML文件
            save_path = filedialog.asksaveasfilename(
                defaultextension=".kml",
                filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")],
                title="保存KML文件"
            )
            
            if save_path:
                tree_string = ET.tostring(kml, encoding='utf-8', method='xml')
                pretty_xml_string = pretty_print_xml(tree_string.decode('utf-8'))
                with open(save_path, "w", encoding="utf-8") as f:
                    f.write(pretty_xml_string)
                messagebox.showinfo("转换成功", f"Excel已成功转换为KML:\n{save_path}")
                if self.update_status:
                    self.update_status(f"Excel转KML完成: {save_path}")
        
        except Exception as e:
            messagebox.showerror("转换失败", f"Excel转KML时发生错误: {e}")
            if self.update_status:
                self.update_status(f"Excel转KML失败: {e}")
    
    def convert_kml_to_excel(self):
        """KML转Excel"""
        file_path = filedialog.askopenfilename(
            title="选择KML文件进行转换",
            filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        
        try:
            points, error_msg = parse_kml_points(file_path)
            if error_msg:
                messagebox.showerror("解析失败", error_msg)
                if self.update_status:
                    self.update_status(f"KML解析失败: {error_msg}")
                return
            
            if not points:
                messagebox.showinfo("无数据", "KML文件中没有找到有效的点数据。")
                if self.update_status:
                    self.update_status("KML文件中无有效点数据")
                return
            
            # 保存Excel文件
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
                title="保存Excel文件"
            )
            
            if save_path:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "KML数据"
                
                # 设置标题行
                sheet["A1"] = "名称"
                sheet["B1"] = "经度"
                sheet["C1"] = "纬度"
                sheet["D1"] = "描述"
                
                # 写入数据
                for row_idx, point in enumerate(points, start=2):
                    sheet[f"A{row_idx}"] = point['name']
                    sheet[f"B{row_idx}"] = point['lon']
                    sheet[f"C{row_idx}"] = point['lat']
                    sheet[f"D{row_idx}"] = point['description']
                
                workbook.save(save_path)
                messagebox.showinfo("转换成功", f"KML已成功转换为Excel:\n{save_path}\n共转换 {len(points)} 个点")
                if self.update_status:
                    self.update_status(f"KML转Excel完成: {len(points)}个点")
        
        except Exception as e:
            messagebox.showerror("转换失败", f"KML转Excel时发生错误: {e}")
            if self.update_status:
                self.update_status(f"KML转Excel失败: {e}")
    
    def convert_address_to_coords_excel(self):
        """地址转经纬度（Excel批量处理）"""
        api_key = self.config.get_amap_api_key()
        if not api_key:
            messagebox.showerror("API Key错误", "请先在配置中设置有效的高德API Key。")
            return
        
        file_path = filedialog.askopenfilename(
            title="选择包含地址的Excel文件",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            results = []
            total_rows = sheet.max_row - 1  # 减去标题行
            processed = 0
            
            if self.update_status:
                self.update_status("正在批量转换地址...")
            
            # 假设第一列是地址，第二列是城市（可选）
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
                if row and row[0]:  # 确保地址不为空
                    address = str(row[0]).strip()
                    city = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    
                    error_msg, lon, lat = get_coords_from_amap(address, city, api_key)
                    
                    if error_msg:
                        results.append({
                            'address': address,
                            'city': city,
                            'lon': 'N/A',
                            'lat': 'N/A',
                            'error': error_msg
                        })
                    else:
                        results.append({
                            'address': address,
                            'city': city,
                            'lon': lon,
                            'lat': lat,
                            'error': ''
                        })
                    
                    processed += 1
                    if self.update_status:
                        self.update_status(f"正在处理地址转换: {processed}/{total_rows}")
            
            # 保存结果
            if results:
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
                    title="保存转换结果"
                )
                
                if save_path:
                    result_workbook = openpyxl.Workbook()
                    result_sheet = result_workbook.active
                    result_sheet.title = "地址转坐标结果"
                    
                    # 设置标题行
                    result_sheet["A1"] = "地址"
                    result_sheet["B1"] = "城市"
                    result_sheet["C1"] = "经度"
                    result_sheet["D1"] = "纬度"
                    result_sheet["E1"] = "错误信息"
                    
                    # 写入结果
                    for row_idx, result in enumerate(results, start=2):
                        result_sheet[f"A{row_idx}"] = result['address']
                        result_sheet[f"B{row_idx}"] = result['city']
                        result_sheet[f"C{row_idx}"] = result['lon']
                        result_sheet[f"D{row_idx}"] = result['lat']
                        result_sheet[f"E{row_idx}"] = result['error']
                    
                    result_workbook.save(save_path)
                    
                    success_count = sum(1 for r in results if not r['error'])
                    messagebox.showinfo("转换完成", 
                                      f"地址转坐标完成！\n"
                                      f"总计: {len(results)} 条\n"
                                      f"成功: {success_count} 条\n"
                                      f"失败: {len(results) - success_count} 条\n"
                                      f"结果已保存到: {save_path}")
                    if self.update_status:
                        self.update_status(f"地址转换完成: {success_count}/{len(results)}")
        
        except Exception as e:
            messagebox.showerror("转换失败", f"地址转坐标时发生错误: {e}")
            if self.update_status:
                self.update_status(f"地址转换失败: {e}")
    
    def convert_coords_to_address_excel(self):
        """经纬度转地址（Excel批量处理）"""
        api_key = self.config.get_amap_api_key()
        if not api_key:
            messagebox.showerror("API Key错误", "请先在配置中设置有效的高德API Key。")
            return
        
        file_path = filedialog.askopenfilename(
            title="选择包含经纬度的Excel文件",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            results = []
            total_rows = sheet.max_row - 1  # 减去标题行
            processed = 0
            
            if self.update_status:
                self.update_status("正在批量转换坐标...")
            
            # 假设第一列是经度，第二列是纬度
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
                if row and len(row) >= 2 and row[0] and row[1]:
                    try:
                        lon = float(row[0])
                        lat = float(row[1])
                        name = str(row[2]) if len(row) > 2 and row[2] else f"点{row_idx}"
                        
                        address = get_address_from_amap(lon, lat, api_key)
                        
                        results.append({
                            'name': name,
                            'lon': lon,
                            'lat': lat,
                            'address': address
                        })
                    except (ValueError, TypeError):
                        results.append({
                            'name': f"点{row_idx}",
                            'lon': 'N/A',
                            'lat': 'N/A',
                            'address': '坐标格式错误'
                        })
                    
                    processed += 1
                    if self.update_status:
                        self.update_status(f"正在处理坐标转换: {processed}/{total_rows}")
            
            # 保存结果
            if results:
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
                    title="保存转换结果"
                )
                
                if save_path:
                    result_workbook = openpyxl.Workbook()
                    result_sheet = result_workbook.active
                    result_sheet.title = "坐标转地址结果"
                    
                    # 设置标题行
                    result_sheet["A1"] = "名称"
                    result_sheet["B1"] = "经度"
                    result_sheet["C1"] = "纬度"
                    result_sheet["D1"] = "地址"
                    
                    # 写入结果
                    for row_idx, result in enumerate(results, start=2):
                        result_sheet[f"A{row_idx}"] = result['name']
                        result_sheet[f"B{row_idx}"] = result['lon']
                        result_sheet[f"C{row_idx}"] = result['lat']
                        result_sheet[f"D{row_idx}"] = result['address']
                    
                    result_workbook.save(save_path)
                    messagebox.showinfo("转换完成", f"坐标转地址完成！\n共处理 {len(results)} 条记录\n结果已保存到: {save_path}")
                    if self.update_status:
                        self.update_status(f"坐标转换完成: {len(results)}条记录")
        
        except Exception as e:
            messagebox.showerror("转换失败", f"坐标转地址时发生错误: {e}")
            if self.update_status:
                self.update_status(f"坐标转换失败: {e}")
    
    def convert_points_to_circles(self):
        """KML点画圆"""
        file_path = filedialog.askopenfilename(
            title="选择包含点的KML文件",
            filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        
        try:
            points, error_msg = parse_kml_points(file_path)
            if error_msg:
                messagebox.showerror("解析失败", error_msg)
                if self.update_status:
                    self.update_status(f"KML解析失败: {error_msg}")
                return
            
            if not points:
                messagebox.showinfo("无数据", "KML文件中没有找到有效的点数据。")
                if self.update_status:
                    self.update_status("KML文件中无有效点数据")
                return
            
            # 询问圆的半径
            radius_str = simpledialog.askstring("设置半径", "请输入圆的半径（米）:", initialvalue="1000")
            if not radius_str:
                return
            
            try:
                radius_meters = float(radius_str)
                if radius_meters <= 0:
                    messagebox.showerror("输入错误", "半径必须大于0")
                    return
            except ValueError:
                messagebox.showerror("输入错误", "请输入有效的数字")
                return
            
            # 创建包含圆形的KML
            kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
            document = ET.SubElement(kml, "Document")
            ET.SubElement(document, "name").text = "点画圆结果"
            
            for point in points:
                circle_placemark = create_kml_circle_placemark(
                    f"{point['name']}_圆形_{radius_meters}m",
                    point['lon'],
                    point['lat'],
                    radius_meters,
                    f"原点: {point['name']}\n半径: {radius_meters}米\n{point['description']}"
                )
                document.append(circle_placemark)
            
            # 保存结果
            save_path = filedialog.asksaveasfilename(
                defaultextension=".kml",
                filetypes=[("KML 文件", "*.kml"), ("所有文件", "*.*")],
                title="保存圆形KML文件"
            )
            
            if save_path:
                tree_string = ET.tostring(kml, encoding='utf-8', method='xml')
                pretty_xml_string = pretty_print_xml(tree_string.decode('utf-8'))
                with open(save_path, "w", encoding="utf-8") as f:
                    f.write(pretty_xml_string)
                messagebox.showinfo("转换成功", 
                                  f"点画圆完成！\n"
                                  f"处理了 {len(points)} 个点\n"
                                  f"半径: {radius_meters} 米\n"
                                  f"结果已保存到: {save_path}")
                if self.update_status:
                    self.update_status(f"点画圆完成: {len(points)}个圆形")
        
        except Exception as e:
            messagebox.showerror("转换失败", f"点画圆时发生错误: {e}")
            if self.update_status:
                self.update_status(f"点画圆失败: {e}")
    
    def download_excel_templates(self):
        """下载Excel模板文件"""
        # 创建模板选择对话框
        template_window = tk.Toplevel(self.parent)
        template_window.title("选择Excel模板")
        template_window.geometry("400x300")
        template_window.transient(self.parent)
        template_window.grab_set()
        
        # 居中显示
        template_window.geometry("+%d+%d" % (
            self.parent.winfo_rootx() + 50,
            self.parent.winfo_rooty() + 50
        ))
        
        main_frame = ttk.Frame(template_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 标题
        title_label = ttk.Label(main_frame, text="选择要下载的Excel模板", 
                               font=("微软雅黑", 12, "bold"))
        title_label.pack(pady=(0, 20))
        
        # 模板按钮
        templates = [
            ("📍 Excel转KML模板", "包含: 名称、经度、纬度、描述列", self.create_excel_to_kml_template),
            ("🏠 地址转坐标模板", "包含: 地址、城市列", self.create_address_to_coords_template),
            ("📌 坐标转地址模板", "包含: 经度、纬度、名称列", self.create_coords_to_address_template)
        ]
        
        for title, desc, command in templates:
            frame = ttk.Frame(main_frame)
            frame.pack(fill=tk.X, pady=5)
            
            btn = ttk.Button(frame, text=title, command=lambda cmd=command: self.download_template(cmd, template_window))
            btn.pack(fill=tk.X)
            
            desc_label = ttk.Label(frame, text=desc, font=("微软雅黑", 8), foreground="gray")
            desc_label.pack()
        
        # 关闭按钮
        close_btn = ttk.Button(main_frame, text="关闭", command=template_window.destroy)
        close_btn.pack(pady=(20, 0))
    
    def download_template(self, template_func, window):
        """下载指定模板"""
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
            title="保存Excel模板"
        )
        
        if save_path:
            try:
                template_func(save_path)
                messagebox.showinfo("下载成功", f"Excel模板已保存到:\n{save_path}")
                if self.update_status:
                    self.update_status(f"模板下载完成: {save_path}")
                window.destroy()
            except Exception as e:
                messagebox.showerror("下载失败", f"保存模板时发生错误: {e}")
                if self.update_status:
                    self.update_status(f"模板下载失败: {e}")
    
    def create_excel_to_kml_template(self, file_path):
        """创建Excel转KML模板"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Excel转KML模板"
        
        # 设置标题行
        headers = ["名称", "经度", "纬度", "描述"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 添加示例数据
        examples = [
            ["示例点1", 116.397428, 39.90923, "这是一个示例点的描述"],
            ["示例点2", 121.473701, 31.230416, "上海市中心"],
            ["示例点3", 113.280637, 23.125178, "广州市中心"]
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                sheet.cell(row=row, column=col, value=value)
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 30
        
        workbook.save(file_path)
    
    def create_address_to_coords_template(self, file_path):
        """创建地址转坐标模板"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "地址转坐标模板"
        
        # 设置标题行
        headers = ["地址", "城市"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 添加示例数据
        examples = [
            ["天安门广场", "北京市"],
            ["外滩", "上海市"],
            ["广州塔", "广州市"],
            ["西湖", "杭州市"]
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                sheet.cell(row=row, column=col, value=value)
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 15
        
        workbook.save(file_path)
    
    def create_coords_to_address_template(self, file_path):
        """创建坐标转地址模板"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "坐标转地址模板"
        
        # 设置标题行
        headers = ["经度", "纬度", "名称"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 添加示例数据
        examples = [
            [116.397428, 39.90923, "北京示例点"],
            [121.473701, 31.230416, "上海示例点"],
            [113.280637, 23.125178, "广州示例点"],
            [120.153576, 30.287459, "杭州示例点"]
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                sheet.cell(row=row, column=col, value=value)
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 15
        
        workbook.save(file_path)
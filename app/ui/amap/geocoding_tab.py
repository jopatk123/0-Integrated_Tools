import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import openpyxl
from ...utils.coordinate_converter import wgs84_to_gcj02
from ...utils import amap_api


class GeocodingTab:
    """批量地理编码选项卡"""
    
    def __init__(self, parent, notebook, theme, config):
        self.parent = parent
        self.notebook = notebook
        self.theme = theme
        self.config = config
        self.excel_file_path = None
        
        # 创建选项卡
        self.frame = ttk.Frame(notebook)
        notebook.add(self.frame, text="📍 批量地理编码")
        self.create_tab()
        
    def create_tab(self):
        """创建批量地理编码选项卡"""
        geocoding_frame = self.frame
        
        # 说明信息
        info_label = tk.Label(geocoding_frame, 
                             text="上传包含经纬度坐标的Excel文件，自动查询对应的行政区域信息\n" +
                                  "Excel格式：第一列为经度(longitude)，第二列为纬度(latitude)，第一行为标题行\n" +
                                  "支持WGS-84坐标系，程序会自动转换为GCJ-02坐标系进行查询",
                             font=("微软雅黑", 9), bg=self.theme.bg_color, fg=self.theme.accent_color,
                             wraplength=600, justify=tk.LEFT)
        info_label.pack(pady=(10, 5), padx=10, anchor=tk.W)
        
        # 文件上传和查询按钮
        btn_frame = tk.Frame(geocoding_frame, bg=self.theme.bg_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 上传文件按钮
        upload_btn = tk.Button(btn_frame, text="上传Excel文件", 
                              command=self.upload_excel,
                              bg=self.theme.button_color, fg="white",
                              font=("微软雅黑", 10), relief=tk.FLAT)
        upload_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 开始查询按钮
        query_btn = tk.Button(btn_frame, text="开始查询", 
                             command=self.start_batch_geocoding,
                             bg=self.theme.button_color, fg="white",
                             font=("微软雅黑", 10), relief=tk.FLAT)
        query_btn.pack(side=tk.LEFT)
        
        # 文件路径显示
        self.file_path_var = tk.StringVar(value="未选择文件")
        file_label = tk.Label(geocoding_frame, textvariable=self.file_path_var,
                             font=("微软雅黑", 9), bg=self.theme.bg_color, fg=self.theme.text_color)
        file_label.pack(pady=5, padx=10, anchor=tk.W)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(geocoding_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, padx=10, pady=5)
        
        # 结果显示
        result_frame = tk.Frame(geocoding_frame, bg=self.theme.bg_color)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tk.Label(result_frame, text="查询结果:", 
                font=("微软雅黑", 10, "bold"), bg=self.theme.bg_color, fg=self.theme.text_color).pack(anchor=tk.W)
        
        text_frame = tk.Frame(result_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.geocoding_result_text = tk.Text(text_frame, height=10, wrap=tk.WORD)
        scrollbar = tk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.geocoding_result_text.yview)
        self.geocoding_result_text.configure(yscrollcommand=scrollbar.set)
        
        self.geocoding_result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def upload_excel(self):
        """上传Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        
        if file_path:
            self.file_path_var.set(f"已选择: {os.path.basename(file_path)}")
            self.excel_file_path = file_path
            self.update_status("文件上传完成")
        
    def start_batch_geocoding(self):
        """开始批量地理编码"""
        if not hasattr(self, 'excel_file_path') or not self.excel_file_path:
            messagebox.showerror("错误", "请先上传Excel文件")
            return
            
        self.update_status("正在进行批量查询...")
        
        # 在新线程中执行批量查询
        thread = threading.Thread(target=self._batch_geocoding_thread)
        thread.daemon = True
        thread.start()
        
    def _batch_geocoding_thread(self):
        """批量地理编码线程"""
        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(self.excel_file_path)
            ws = wb.active
            
            # 获取数据行数，找到实际有数据的行
            max_row = ws.max_row
            data_rows = []
            
            # 先扫描所有行，找出有效的数据行
            for row in range(2, max_row + 1):  # 从第2行开始检查
                lng = ws.cell(row=row, column=1).value
                lat = ws.cell(row=row, column=2).value
                
                # 检查是否为有效的数值坐标
                if (lng is not None and lat is not None and 
                    isinstance(lng, (int, float)) and isinstance(lat, (int, float)) and
                    -180 <= lng <= 180 and -90 <= lat <= 90):
                    data_rows.append(row)
            
            if not data_rows:
                self.parent.after(0, lambda: messagebox.showerror("错误", "未找到有效的坐标数据，请检查Excel文件格式"))
                return
            
            results = []
            total_rows = len(data_rows)
            
            for i, row in enumerate(data_rows):
                lng = ws.cell(row=row, column=1).value
                lat = ws.cell(row=row, column=2).value
                
                try:
                    # 将WGS-84坐标转换为GCJ-02坐标（高德地图使用的坐标系）
                    lng_gcj, lat_gcj = wgs84_to_gcj02(float(lng), float(lat))
                    
                    # 直接调用高德API进行逆地理编码
                    api_key = self.config.get_amap_api_key()
                    if not api_key:
                        raise Exception("未配置API密钥")
                    
                    result = amap_api.regeocode(lng_gcj, lat_gcj)
                    
                    # 在结果中记录坐标转换信息
                    coord_conversion = f"坐标转换: WGS-84({lng:.6f}, {lat:.6f}) -> GCJ-02({lng_gcj:.6f}, {lat_gcj:.6f})"
                    
                    # 直接从API结果中提取地址信息
                    if result['status'] == 'success':
                        province = result.get('province', '')
                        city = result.get('city', '')
                        district = result.get('district', '')
                        formatted_address = result.get('formatted_address', '')
                        
                        # 更新Excel单元格
                        ws.cell(row=row, column=3, value=province)
                        ws.cell(row=row, column=4, value=city)
                        ws.cell(row=row, column=5, value=district)
                        ws.cell(row=row, column=6, value=formatted_address)
                        
                        results.append(f"行{row}: {coord_conversion} -> 查询成功: {province} {city} {district}")
                    else:
                        # API调用失败，记录错误信息
                        error_msg = result.get('message', '未知错误')
                        ws.cell(row=row, column=3, value="查询失败")
                        ws.cell(row=row, column=4, value="查询失败")
                        ws.cell(row=row, column=5, value="查询失败")
                        ws.cell(row=row, column=6, value=f"API错误: {error_msg}")
                        
                        results.append(f"行{row}: {coord_conversion} -> 查询失败: {error_msg}")
                        
                except Exception as e:
                    # 处理异常
                    ws.cell(row=row, column=3, value="查询失败")
                    ws.cell(row=row, column=4, value="查询失败")
                    ws.cell(row=row, column=5, value="查询失败")
                    ws.cell(row=row, column=6, value=f"错误: {str(e)}")
                    
                    results.append(f"行{row}: 查询失败 - {str(e)}")
                
                # 更新进度
                progress = ((i + 1) / total_rows) * 100
                self.parent.after(0, lambda p=progress: self.progress_var.set(p))
                
                # 更新结果显示
                if (i + 1) % 10 == 0 or i == total_rows - 1:  # 每10条或最后一条更新一次显示
                    self.parent.after(0, lambda r=results.copy(): self._update_result_display(r))
            
            # 保存结果到新文件
            self.parent.after(0, lambda: self._save_results(wb, results))
            
        except Exception as e:
            self.parent.after(0, lambda: messagebox.showerror("错误", f"批量查询失败: {str(e)}"))
            self.parent.after(0, lambda: self.update_status("查询失败"))
    
    def _update_result_display(self, results):
        """更新结果显示"""
        self.geocoding_result_text.delete(1.0, tk.END)
        for result in results:
            self.geocoding_result_text.insert(tk.END, result + "\n")
        self.geocoding_result_text.see(tk.END)
    
    def _save_results(self, wb, results):
        """保存查询结果"""
        try:
            # 生成输出文件名
            base_name = os.path.splitext(self.excel_file_path)[0]
            output_file = f"{base_name}_行政区域查询结果.xlsx"
            
            # 保存Excel文件
            wb.save(output_file)
            
            # 更新状态和结果显示
            self.update_status(f"查询完成，结果已保存到: {os.path.basename(output_file)}")
            
            # 显示完整结果
            self._update_result_display(results)
            
            # 重置进度条
            self.progress_var.set(0)
            
            messagebox.showinfo("完成", f"批量查询完成！\n结果已保存到: {output_file}")
            
        except Exception as e:
            messagebox.showerror("错误", f"保存结果失败: {str(e)}")
    
    def update_status(self, message):
        """更新状态信息"""
        # 这里可以添加状态栏更新逻辑
        # 如果父组件有状态栏，可以通过回调更新
        pass